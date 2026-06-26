"""order_export.py —— 订单 xlsx 导出层(从 main.py 原样搬出)。

取数与格式分离:本模块只负责"给定已取好的 records + 导出目录 → 生成 xlsx 文件",
不碰数据层、不读 env、不 import main。EXPORT_DIR 由 main 按铁律(e)注入(export_dir 参数),
records 由调用方(main 的 collect_order_records 适配器)取好后传入。

collect_order_records 取数适配器仍在 main(依赖数据层 query_order_payloads,属 P5),
P5 抽 store_sqlite 时再一并迁移。
"""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ORDER_SUMMARY_HEADERS = [
    "门店/区域",
    "商品",
    "单位",
    "数量合计",
    "订单行数",
    "最近创建时间",
]

ORDER_CONTRACT_EXPORT_HEADERS = [
    "ID",
    "类型",
    "来源",
    "状态",
    "已确认",
    "门店/区域",
    "订单号",
    "下单人",
    "下单日期",
    "送达日期",
    "变更类型",
    "行号",
    "商品编码",
    "商品名称",
    "规格",
    "单位",
    "数量",
    "单价",
    "分类",
    "原始文本",
    "原始引用",
    "创建时间",
]


def order_record_to_export_row(record: dict[str, str]) -> list[str]:
    return [
        record["id"],
        record["kind"],
        record["source"],
        record["status"],
        record["confirmed"],
        record["store"],
        record["order_no"],
        record["orderer"],
        record["order_date"],
        record["deliver_date"],
        record["change_type"],
        record["line_no"],
        record["code"],
        record["name"],
        record["spec"],
        record["unit"],
        record["qty"],
        record["price"],
        record["category"],
        record["raw_text"],
        record["raw_ref"],
        record["created_at"],
    ]


def parse_quantity_number(value: str) -> float | None:
    match = re.search(r"-?\d+(?:\.\d+)?", value)
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def format_quantity_total(value: float) -> str:
    if value.is_integer():
        return str(int(value))
    return f"{value:.2f}".rstrip("0").rstrip(".")


def build_order_summary_rows(records: list[dict[str, str]]) -> list[list[str]]:
    groups: dict[tuple[str, str, str], dict[str, Any]] = {}
    for record in records:
        key = (record["store"], record["name"], record["unit"])
        group = groups.setdefault(
            key,
            {
                "total": 0.0,
                "numeric_count": 0,
                "raw_quantities": [],
                "count": 0,
                "latest_created_at": "",
            },
        )
        group["count"] += 1
        quantity = record["qty"]
        number = parse_quantity_number(quantity)
        if number is None:
            if quantity:
                group["raw_quantities"].append(quantity)
        else:
            group["total"] += number
            group["numeric_count"] += 1
        if record["created_at"] > group["latest_created_at"]:
            group["latest_created_at"] = record["created_at"]

    rows: list[list[str]] = []
    for (store, product, unit), group in sorted(groups.items()):
        quantity_parts: list[str] = []
        if group["numeric_count"]:
            quantity_parts.append(format_quantity_total(float(group["total"])))
        if group["raw_quantities"]:
            quantity_parts.append("、".join(group["raw_quantities"]))
        rows.append(
            [
                store,
                product,
                unit,
                "；".join(quantity_parts),
                str(group["count"]),
                str(group["latest_created_at"]),
            ]
        )
    return rows


def write_order_table_sheet(sheet, headers: list[str], rows: list[list[str]], widths: list[int]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="2F5597")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def build_order_export(records: list[dict[str, str]], export_dir: Path) -> Path:
    workbook = Workbook()
    detail_sheet = workbook.active
    detail_sheet.title = "全部订单"
    detail_rows = [order_record_to_export_row(record) for record in records]
    write_order_table_sheet(
        detail_sheet,
        ORDER_CONTRACT_EXPORT_HEADERS,
        detail_rows,
        [10, 10, 10, 10, 10, 22, 18, 14, 14, 14, 12, 8, 16, 28, 20, 10, 12, 12, 14, 36, 44, 20],
    )

    summary_sheet = workbook.create_sheet("按门店商品汇总")
    write_order_table_sheet(
        summary_sheet,
        ORDER_SUMMARY_HEADERS,
        build_order_summary_rows(records),
        [24, 26, 10, 16, 12, 20],
    )

    export_dir.mkdir(parents=True, exist_ok=True)
    output_path = export_dir / f"orders-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
    workbook.save(output_path)
    return output_path


__all__ = [
    "ORDER_SUMMARY_HEADERS",
    "ORDER_CONTRACT_EXPORT_HEADERS",
    "order_record_to_export_row",
    "parse_quantity_number",
    "format_quantity_total",
    "build_order_summary_rows",
    "write_order_table_sheet",
    "build_order_export",
]
