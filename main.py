import asyncio
import base64
import hashlib
import hmac
import json
import logging
import mimetypes
import os
import re
import socket
import sqlite3
import struct
import sys
import time
import traceback
import xml.etree.ElementTree as ET
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from threading import Lock
from typing import Any
from urllib.parse import unquote

from dotenv import load_dotenv
from fastapi import BackgroundTasks, FastAPI, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, PlainTextResponse
from Crypto.Cipher import AES
import httpx
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.datetime import from_excel
from openpyxl.utils import get_column_letter
from openai import OpenAI
from pydantic import BaseModel, Field
from wx_crypt import WXBizMsgCrypt, WxChannel_Wecom

from services.business_intent import (
    INTENT_CANCEL,
    INTENT_CHAT,
    INTENT_CONFIRM,
    INTENT_EXIT,
    INTENT_MODIFY,
    INTENT_REJECT,
    INTENT_UNCLEAR,
    BusinessIntent,
    classify_business_intent,
)
import agent_router
import models
import store_sqlite  # 本地 SQLite 回退后端（双轨的 SQLite 这一轨；lock/db_file 由 main 注入）
from order_normalize import *  # 订单归一化纯函数层（门面 re-export，调用点/测试可见性不变）
from excel_import import *  # Excel 订单解析层（门面 re-export）
from llm_json import *  # 大模型输出 JSON 提取通用 helper（门面 re-export）
from vision_import import *  # 视觉/照片解析层（client/model 由 main 注入；门面 re-export）
from order_export import *  # 订单 xlsx 导出层（records/export_dir 由 main 注入；门面 re-export）
from store_sqlite import summarize_order_for_reply  # 撤回回复模板（models 分支格式化用；SQLite 分支在 store_sqlite 内自用）
from receipt_logic import *  # 入库(receipt)领域：RECEIPT 常量 + 归一化/missing/summarize/status helper（门面 re-export）

load_dotenv()

DEFAULT_LLM_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DEFAULT_MODEL_NAME = "qwen3-vl-plus"
DEFAULT_MAX_HISTORY_MESSAGES = 20
DEFAULT_SYSTEM_PROMPT = "你是运行在微信里的公司客服助手。说话自然、简洁，像真人客服；客户闲聊可以接一两句，但不要把客服窗口变成闲聊室。涉及价格、交期、投诉、纠纷、退款、发票等事项，不要瞎承诺，明确转人工处理。"
CUSTOMER_CHAT_PROMPT = "你是运行在微信里的公司客服。普通聊天要自然、简短，像真人客服。不要进入需求访谈，不要问客户工作流程、审批、排班、数据整理、痛点、频率，也不要说自己只有需求访谈模式。客户问订单/入库怎么用时，只给简短操作提示。涉及价格、交期、投诉、纠纷、退款、发票等事项，不要承诺，回复转人工处理。不要声称自己正在入库、同步、查询数据库、生成单号或保存成功；这些业务动作只能由程序命令层完成。"
DEFAULT_WECOM_BOT_NAME = "食品厂机器人"
DEFAULT_WECOM_KF_SYNC_LIMIT = 100
DEFAULT_HTTP_TIMEOUT_SECONDS = 20
DEFAULT_EXPORT_DIR = "exports"
DEFAULT_SESSION_STATE_FILE = "session_state.json"
DEFAULT_ORDER_DB_FILE = "orders.db"
DEFAULT_RECEIPT_DB_FILE = "receipts.db"
DEFAULT_VISION_BASE_URL = "https://dashscope.aliyuncs.com/compatible-mode/v1"
DEFAULT_VISION_MODEL = "qwen3-vl-plus"
# EXCEL_MAX_SCAN_ROWS / EXCEL_MAX_SCAN_COLUMNS 已移至 excel_import.py（门面 re-export）。
APP_BUILD_LABEL = os.getenv("APP_BUILD_LABEL", "excel-diagnostics-2026-06-24")


def get_int_env(name: str, default: int) -> int:
    value = os.getenv(name)
    if value is None:
        return default

    try:
        parsed = int(value)
    except ValueError:
        logging.warning("Invalid %s=%s, using default=%s", name, value, default)
        return default

    if parsed < 1:
        logging.warning("Invalid %s=%s, using default=%s", name, value, default)
        return default

    return parsed


def get_bool_env(name: str, default: bool) -> bool:
    value = os.getenv(name)
    if value is None:
        return default
    return value.strip().lower() in {"1", "true", "yes", "on"}


def load_system_prompt() -> str:
    prompt_file = os.getenv("PROMPT_FILE")
    if prompt_file:
        try:
            return Path(prompt_file).read_text(encoding="utf-8").strip()
        except OSError as exc:
            logging.warning("Failed to read PROMPT_FILE=%s: %s", prompt_file, exc)

    return os.getenv("SYSTEM_PROMPT") or DEFAULT_SYSTEM_PROMPT


LOG_FORMAT = "%(asctime)s %(levelname)s %(name)s %(message)s"
logging.basicConfig(
    level=os.getenv("LOG_LEVEL", "INFO"),
    format=LOG_FORMAT,
    stream=sys.stdout,
    force=True,
)
logger = logging.getLogger("wechatclaw")
logger.setLevel(os.getenv("LOG_LEVEL", "INFO"))

log_file = os.getenv("LOG_FILE")
if log_file:
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setFormatter(logging.Formatter(LOG_FORMAT))
    logger.addHandler(file_handler)

app = FastAPI(title="WechatClaw Phase 0.5 AI Backend MVP")

LLM_API_KEY = os.getenv("LLM_API_KEY") or os.getenv("OPENAI_API_KEY")
LLM_BASE_URL = os.getenv("LLM_BASE_URL", DEFAULT_LLM_BASE_URL)
MODEL_NAME = os.getenv("MODEL_NAME") or DEFAULT_MODEL_NAME
MAX_HISTORY_MESSAGES = get_int_env("MAX_HISTORY_MESSAGES", DEFAULT_MAX_HISTORY_MESSAGES)
PROMPT_FILE = os.getenv("PROMPT_FILE")
SYSTEM_PROMPT = load_system_prompt()
PROMPT_TURN_CONTEXT = get_bool_env("PROMPT_TURN_CONTEXT", bool(PROMPT_FILE))
MEMORY_FILE = Path(os.getenv("MEMORY_FILE", "memory.json"))
SESSION_STATE_FILE = Path(os.getenv("SESSION_STATE_FILE", DEFAULT_SESSION_STATE_FILE))
ORDER_DB_FILE = Path(os.getenv("ORDER_DB_FILE", DEFAULT_ORDER_DB_FILE))
RECEIPT_DB_FILE = Path(os.getenv("RECEIPT_DB_FILE", DEFAULT_RECEIPT_DB_FILE))
VISION_API_KEY = os.getenv("VISION_API_KEY") or os.getenv("DASHSCOPE_API_KEY", "")
VISION_BASE_URL = os.getenv("VISION_BASE_URL", DEFAULT_VISION_BASE_URL)
VISION_MODEL = os.getenv("VISION_MODEL", DEFAULT_VISION_MODEL)
ROBOT_API_TOKEN = os.getenv("ROBOT_API_TOKEN", "")
PUBLIC_BASE_URL = os.getenv("PUBLIC_BASE_URL", "").rstrip("/")
MEMORY_LOCK = Lock()
WECOM_CALLBACK_TOKEN = os.getenv("WECOM_CALLBACK_TOKEN") or os.getenv("WX_BOT_TOKEN")
WECOM_ENCODING_AES_KEY = os.getenv("WECOM_ENCODING_AES_KEY") or os.getenv("WX_BOT_AES_KEY")
WECOM_CORP_ID = os.getenv("WECOM_CORP_ID") or os.getenv("WX_BOT_CORP_ID", "")
WECOM_BOT_NAME = os.getenv("WECOM_BOT_NAME", DEFAULT_WECOM_BOT_NAME)
WECOM_KF_CORP_ID = os.getenv("WECOM_KF_CORP_ID") or WECOM_CORP_ID
WECOM_KF_SECRET = os.getenv("WECOM_KF_SECRET")
WECOM_KF_CALLBACK_TOKEN = os.getenv("WECOM_KF_CALLBACK_TOKEN")
WECOM_KF_ENCODING_AES_KEY = os.getenv("WECOM_KF_ENCODING_AES_KEY")
WECOM_KF_CURSOR_FILE = Path(os.getenv("WECOM_KF_CURSOR_FILE", "kf_cursors.json"))
WECOM_KF_SYNC_LIMIT = get_int_env("WECOM_KF_SYNC_LIMIT", DEFAULT_WECOM_KF_SYNC_LIMIT)
HTTP_TIMEOUT_SECONDS = get_int_env("HTTP_TIMEOUT_SECONDS", DEFAULT_HTTP_TIMEOUT_SECONDS)
EXPORT_DIR = Path(os.getenv("EXPORT_DIR", DEFAULT_EXPORT_DIR))
EXPORT_TOKEN = os.getenv("EXPORT_TOKEN")
SEEN_WECOM_MSG_IDS: set[str] = set()
SEEN_WECOM_MSG_IDS_LOCK = Lock()
SEEN_WECOM_KF_MSG_IDS: set[str] = set()
SEEN_WECOM_KF_MSG_IDS_LOCK = Lock()
WECOM_KF_CURSOR_LOCK = Lock()
WECOM_KF_ACCESS_TOKEN_LOCK = Lock()
SESSION_STATE_LOCK = Lock()
ORDER_LOCK = Lock()
ORDER_DB_LOCK = Lock()
RECEIPT_DB_LOCK = Lock()
WECOM_KF_ACCESS_TOKEN = ""
WECOM_KF_ACCESS_TOKEN_EXPIRES_AT = 0.0

if not LLM_API_KEY:
    raise RuntimeError("LLM_API_KEY is missing. Check your .env file.")

client = OpenAI(
    api_key=LLM_API_KEY,
    base_url=LLM_BASE_URL,
)
vision_client = OpenAI(
    api_key=VISION_API_KEY,
    base_url=VISION_BASE_URL,
) if VISION_API_KEY else None


class ChatRequest(BaseModel):
    user_id: str = Field(..., min_length=1)
    message: str = Field(..., min_length=1)


class ChatResponse(BaseModel):
    user_id: str
    answer: str
    history_length: int


class MemoryLengthResponse(BaseModel):
    user_id: str
    history_length: int


class DeleteMemoryResponse(BaseModel):
    deleted: bool
    user_id: str


class WecomMessage(BaseModel):
    msg_type: str
    chat_type: str
    chat_id: str
    msg_id: str
    sender_user_id: str
    sender_name: str
    content: str


class WecomKfEvent(BaseModel):
    token: str
    open_kfid: str
    event: str
    create_time: str


class WecomKfApiError(RuntimeError):
    def __init__(self, path: str, data: dict[str, Any]) -> None:
        self.path = path
        self.data = data
        self.errcode = data.get("errcode")
        super().__init__(f"WeCom API {path} failed: {data}")


class MarkFetchedRequest(BaseModel):
    ids: list[int] = Field(default_factory=list)


class IdsRequest(BaseModel):
    ids: list[Any] = Field(default_factory=list)


class TextOrderImportRequest(BaseModel):
    user_id: str = Field(..., min_length=1)
    message: str = Field(..., min_length=1)
    confirm: bool = False
    raw_ref: str | None = None


def load_memory() -> dict:
    if models.is_enabled():
        return models.load_memory()

    if not MEMORY_FILE.exists():
        return {}

    raw_memory = MEMORY_FILE.read_text(encoding="utf-8").strip()
    if not raw_memory:
        return {}

    try:
        data = json.loads(raw_memory)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="memory.json is not valid JSON") from exc

    if not isinstance(data, dict):
        raise HTTPException(status_code=500, detail="memory.json must contain a JSON object")

    return data


def save_memory(memory: dict) -> None:
    if models.is_enabled():
        models.save_memory(memory)
        return

    MEMORY_FILE.write_text(
        json.dumps(memory, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def load_session_state() -> dict[str, dict[str, Any]]:
    if models.is_enabled():
        return models.load_session_state()

    if not SESSION_STATE_FILE.exists():
        return {}

    raw_state = SESSION_STATE_FILE.read_text(encoding="utf-8").strip()
    if not raw_state:
        return {}

    try:
        data = json.loads(raw_state)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail="session_state.json is not valid JSON") from exc

    if not isinstance(data, dict):
        raise HTTPException(status_code=500, detail="session_state.json must contain a JSON object")

    state: dict[str, dict[str, Any]] = {}
    for user_id, record in data.items():
        if isinstance(record, dict):
            state[str(user_id)] = record
    return state


def save_session_state(state: dict[str, dict[str, Any]]) -> None:
    if models.is_enabled():
        models.save_session_state(state)
        return

    SESSION_STATE_FILE.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def insert_order_payload(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = normalize_order_payload(payload)
    if normalized.get("confirmed"):
        missing = order_draft_missing_fields(normalized)
        if missing:
            raise ValueError("confirmed order missing fields: " + ",".join(missing))

    if models.is_enabled():
        return models.insert_order_payload(normalized)

    return store_sqlite.insert_order_payload(normalized, db_file=ORDER_DB_FILE, lock=ORDER_DB_LOCK)


def query_order_payloads(
    status: str | None = None,
    ids: list[int] | None = None,
    order_date: str | None = None,
) -> list[dict[str, Any]]:
    if models.is_enabled():
        return models.query_order_payloads(status=status, ids=ids, order_date=order_date)

    return store_sqlite.query_order_payloads(
        db_file=ORDER_DB_FILE, status=status, ids=ids, order_date=order_date
    )


def mark_order_payloads_fetched(ids: list[int]) -> dict[str, list[int]]:
    if models.is_enabled():
        return models.mark_order_payloads_fetched(ids)

    return store_sqlite.mark_order_payloads_fetched(ids, db_file=ORDER_DB_FILE, lock=ORDER_DB_LOCK)


def unmark_order_payloads(ids: list[int]) -> dict[str, list[int]]:
    if models.is_enabled():
        return models.unmark_order_payloads(ids)

    return store_sqlite.unmark_order_payloads(ids, db_file=ORDER_DB_FILE, lock=ORDER_DB_LOCK)


def cancel_latest_order_for_user(user_id: str) -> str:
    if models.is_enabled():
        result = models.cancel_latest_order_for_user(user_id)
        if result.get("outcome") == "not_found":
            return "没找到你最近确认的订单，暂时没有可撤回的。"
        if result.get("outcome") == "fetched":
            return "这单已被排产/发货使用，不能直接撤回，需要联系数据部处理。"
        payload = result.get("payload") if isinstance(result.get("payload"), dict) else {}
        return f"好，刚那单（{summarize_order_for_reply(payload)}）撤回了，重新发我吧。"

    return store_sqlite.cancel_latest_order_for_user(user_id, db_file=ORDER_DB_FILE, lock=ORDER_DB_LOCK)


def cancel_latest_receipt_for_user(user_id: str) -> str:
    today = datetime.now().date().isoformat()
    if models.is_enabled():
        result = models.cancel_latest_receipt_for_user(user_id, today)
        if result.get("outcome") == "cancelled":
            payload = result.get("payload") if isinstance(result.get("payload"), dict) else {}
            return f"好，刚那条入库记录（{summarize_receipt_for_reply(payload)}）撤回了。"
        if result.get("outcome") == "fetched":
            return "这条入库记录已被入库工具使用，不能直接撤回，需要联系数据部处理。"
        return "没找到你今天确认的入库记录，暂时没有可撤回的。"

    return store_sqlite.cancel_latest_receipt_for_user(user_id, today, db_file=RECEIPT_DB_FILE, lock=RECEIPT_DB_LOCK)


def order_draft_has_content(draft: dict[str, Any]) -> bool:
    if not isinstance(draft, dict):
        return False
    if draft.get("store"):
        return True
    items = draft.get("items") if isinstance(draft.get("items"), list) else []
    return any(
        isinstance(item, dict) and (item.get("name") or item.get("qty") is not None)
        for item in items
    )


def receipt_draft_has_content(draft: dict[str, Any]) -> bool:
    if not isinstance(draft, dict):
        return False
    items = draft.get("items") if isinstance(draft.get("items"), list) else []
    return any(
        isinstance(item, dict) and (item.get("name") or item.get("qty") is not None)
        for item in items
    )


def missing_fields_reply(missing: list[str], *, receipt: bool = False) -> str:
    if not missing:
        return ""
    subject = "这条入库记录" if receipt else "这单"
    return f"{subject}还差：{'、'.join(missing)}。补我一下就行，发“取消”可以清空草稿。"


def insert_receipt_payload(payload: dict[str, Any]) -> dict[str, Any]:
    normalized = normalize_receipt_payload(payload)
    missing = receipt_missing_fields(normalized)
    if missing:
        raise ValueError("receipt missing fields: " + ",".join(missing))

    if models.is_enabled():
        return models.insert_receipt_payload(normalized)

    return store_sqlite.insert_receipt_payload(normalized, db_file=RECEIPT_DB_FILE, lock=RECEIPT_DB_LOCK)


def query_receipt_payloads(date: str) -> list[dict[str, Any]]:
    return query_receipt_payloads_by_status(date, RECEIPT_STATUS_NEW)


def update_receipt_payload_status(
    ids: list[Any],
    target_status: str,
) -> dict[str, list[str]]:
    if models.is_enabled():
        if target_status == RECEIPT_STATUS_FETCHED:
            return models.mark_receipt_payloads_fetched(ids)
        return models.unmark_receipt_payloads(ids)

    return store_sqlite.update_receipt_payload_status(ids, target_status, db_file=RECEIPT_DB_FILE, lock=RECEIPT_DB_LOCK)


def mark_receipt_payloads_fetched(ids: list[Any]) -> dict[str, list[str]]:
    return update_receipt_payload_status(ids, RECEIPT_STATUS_FETCHED)


def unmark_receipt_payloads(ids: list[Any]) -> dict[str, list[str]]:
    return update_receipt_payload_status(ids, RECEIPT_STATUS_CONFIRMED)


def query_receipt_payloads_by_status(date: str, status: str | None = None) -> list[dict[str, Any]]:
    if models.is_enabled():
        return models.query_receipt_payloads(date, status=status)

    return store_sqlite.query_receipt_payloads_by_status(date, status, db_file=RECEIPT_DB_FILE)


SESSION_MODE_CHAT = "chat"
SESSION_MODE_ORDER = "order"
SESSION_MODE_RECEIPT = "receipt"
SESSION_MODES = {SESSION_MODE_CHAT, SESSION_MODE_ORDER, SESSION_MODE_RECEIPT}

ORDER_MODE_COMMANDS = {"订单", "录单", "下单", "订单模式", "开始订单", "开始录单"}
RECEIPT_MODE_COMMANDS = {"入库", "入库模式", "产成品入库", "开始入库", "成品入库"}
EXIT_MODE_COMMANDS = {"退出", "结束", "不弄了", "算了", "返回", "退出订单", "退出入库", "结束订单", "结束入库"}
STATUS_COMMANDS = {"状态", "我在哪", "我在哪儿", "当前状态", "现在状态", "现在是什么模式"}
MODE_HELP_COMMANDS = {"模式", "有哪些模式", "有什么模式", "你有哪些模式", "你有什么模式", "怎么用", "你会什么", "功能", "帮助"}
REVOKE_COMMANDS = {
    "撤回",
    "撤销",
    "撤回上一单",
    "撤销上一单",
    "撤回刚刚的入库",
    "撤销刚刚的入库",
    "撤回入库",
    "撤销入库",
    "删了上一单",
    "删了",
    "刚那个不对",
    "刚才那个不对",
    "上一单不对",
}
ORDER_EXPORT_COMMANDS = {"导出订单", "订单导出", "下载订单", "订单表", "导出订单表"}
ORDER_CONFIRM_COMMANDS = {"确认", "确认订单", "保存", "保存订单", "提交", "提交订单"}
CONFIRM_LIKE_KEYWORDS = {"确认", "确认无误", "没问题", "可以", "对的", "是的", "保存", "提交", "录入", "写库", "入数据库", "直接入库", "记下"}
ORDER_STORAGE_QUERY_KEYWORDS = {"入库结果", "同步结果", "订单库", "数据库", "拉订单库", "同步订单", "查订单", "查一下订单", "看一下入库"}
ORDER_CANCEL_COMMANDS = {"取消", "取消订单", "取消草稿", "清空", "清空订单", "清空草稿", "不要了"}
ORDER_DRAFT_VIEW_COMMANDS = {
    "当前订单",
    "订单草稿",
    "查看当前订单",
    "看当前订单",
    "看看当前订单",
    "查看订单草稿",
    "看订单草稿",
    "查看草稿",
    "看草稿",
}
ORDER_DRAFT_VIEW_KEYWORDS = {
    "当前订单",
    "订单草稿",
    "订单内容",
    "当前草稿",
    "看看这单",
    "看这单",
    "查看这单",
    "这单有啥",
    "这单有什么",
    "这张订单",
    "重复一遍订单",
    "重复订单",
    "复述订单",
    "再说一遍订单",
    "再发一遍订单",
    "订单再发一遍",
}
ORDER_QUERY_KEYWORDS = {"查", "查询", "看", "结果", "同步", "拉取", "有没有", "了吗", "是否", "状态"}
BUSINESS_NEGATION_KEYWORDS = {"不要", "不用", "别", "先别", "不需要", "取消", "撤回", "退"}
QUESTION_LIKE_KEYWORDS = {"吗", "么", "?", "？", "能不能", "可不可以", "是否", "怎么", "如何", "什么", "多少", "几号", "价格", "发票"}
SOFT_CONFIRM_COMMANDS = {
    "ok",
    "okay",
    "yes",
    "y",
    "可以",
    "可以的",
    "行",
    "行的",
    "好",
    "好的",
    "对",
    "对的",
    "是",
    "是的",
    "没错",
    "没问题",
    "确认无误",
    "记下",
    "录入",
    "写库",
    "入数据库",
    "直接入库",
}
GLOBAL_ROUTE_CHAT = "chat"
GLOBAL_ROUTE_ORDER_TEXT = "order_text"
GLOBAL_ROUTE_ENTER_ORDER = "enter_order"
GLOBAL_ROUTE_ENTER_RECEIPT = "enter_receipt"
GLOBAL_ROUTE_ORDER_QUERY = "order_query"
GLOBAL_ROUTE_UNCLEAR = "unclear"

# 订单领域常量（ORDER_KIND_*/ORDER_SOURCE_*/ORDER_STATUS_*/ORDER_CHANGE_*/
# ORDER_KINDS/ORDER_SOURCES/ORDER_STATUSES/ORDER_CHANGE_TYPES/BASE_ORDER_FIELDS/
# PATCH_ORDER_FIELDS）已移至 order_normalize.py，经 `from order_normalize import *` re-export。
# RECEIPT_STATUS_* / RECEIPT_API_STATUSES / RECEIPT_STORAGE_STATUSES 已移至 receipt_logic.py（门面 re-export）。

BASE_ITEM_FIELDS = [
    "code",
    "name",
    "spec",
    "unit",
    "qty",
    "price",
    "category",
]
PATCH_ITEM_FIELDS = [
    "code",
    "name",
    "spec",
    "unit",
    "qty",
]

# ORDER_SUMMARY_HEADERS / ORDER_CONTRACT_EXPORT_HEADERS 已移至 order_export.py（门面 re-export）。
# EXCEL_HEADER_ALIASES / EXCEL_METADATA_LABELS 已移至 excel_import.py（门面 re-export）。


def normalize_command(message: str) -> str:
    return re.sub(r"\s+", "", message.strip()).lower()


def command_contains_any(command: str, keywords: set[str]) -> bool:
    return any(keyword and keyword in command for keyword in keywords)


def is_exit_mode_command(command: str) -> bool:
    return command in EXIT_MODE_COMMANDS or command_contains_any(
        command,
        {"不弄", "算了", "退出", "返回普通", "结束订单", "结束入库"},
    )


def is_revoke_command(command: str) -> bool:
    return command in REVOKE_COMMANDS or command_contains_any(
        command,
        {"撤回", "撤销", "删了", "删除上一", "刚那个不对", "刚才那个不对", "上一单不对"},
    )


def is_receipt_revoke_target(command: str) -> bool:
    return command_contains_any(command, {"入库", "入库记录", "成品", "产成品"})


def is_status_command(command: str) -> bool:
    return command in STATUS_COMMANDS


def is_business_query_or_negated(command: str) -> bool:
    return command_contains_any(command, ORDER_QUERY_KEYWORDS | BUSINESS_NEGATION_KEYWORDS)


def is_order_mode_command(command: str) -> bool:
    if command in ORDER_MODE_COMMANDS:
        return True
    if is_business_query_or_negated(command):
        return False
    if command_contains_any(command, {"我要下单", "帮我下单", "要下单", "我要录单", "帮我录单", "要录单", "我要录一单", "帮我录一单", "录一单"}):
        return True
    if "订单" in command and command_contains_any(command, {"发订单", "传订单", "录订单", "下订单", "订单表", "订单图片", "订单照片", "订单模式", "开始订单"}):
        return True
    return False


def is_receipt_mode_command(command: str) -> bool:
    if command in RECEIPT_MODE_COMMANDS:
        return True
    if is_business_query_or_negated(command) or command_contains_any(command, QUESTION_LIKE_KEYWORDS):
        return False
    if "入库" in command and command_contains_any(command, {"开始", "发", "传", "照片", "图片", "模式", "产成品", "成品", "录", "记"}):
        return True
    return False


def is_mode_help_command(command: str) -> bool:
    if command in MODE_HELP_COMMANDS:
        return True
    return "模式" in command and command_contains_any(command, {"哪些", "什么", "有啥", "怎么", "功能"})


def is_question_like_command(command: str) -> bool:
    return command_contains_any(command, QUESTION_LIKE_KEYWORDS)


def is_confirm_command(command: str, *, has_draft: bool = True) -> bool:
    if is_question_like_command(command):
        return False
    if command in ORDER_CONFIRM_COMMANDS:
        return True
    if command_contains_any(command, {"取消", "撤回", "不对", "不是", "别", "不要"}):
        return False
    if not has_draft:
        return False
    if command in SOFT_CONFIRM_COMMANDS:
        return True
    return False


def is_order_storage_query_command(command: str) -> bool:
    return command_contains_any(command, ORDER_STORAGE_QUERY_KEYWORDS)


def is_order_draft_view_command(command: str) -> bool:
    return command in ORDER_DRAFT_VIEW_COMMANDS or command_contains_any(command, ORDER_DRAFT_VIEW_KEYWORDS)


def mode_display_name(mode: str) -> str:
    if mode == SESSION_MODE_ORDER:
        return "订单模式"
    if mode == SESSION_MODE_RECEIPT:
        return "入库模式"
    return "普通聊天模式"


def draft_mode_display_name(mode: str) -> str:
    if mode == SESSION_MODE_ORDER:
        return "订单草稿"
    if mode == SESSION_MODE_RECEIPT:
        return "入库草稿"
    return "业务草稿"


def has_raw_order_draft(record: dict[str, Any]) -> bool:
    draft = record.get("order_draft")
    return isinstance(draft, dict) and order_draft_has_content(normalize_order_draft(draft))


def has_raw_receipt_draft(record: dict[str, Any]) -> bool:
    draft = record.get("receipt_draft")
    return isinstance(draft, dict) and receipt_draft_has_content(normalize_receipt_payload(draft))


def conflicting_draft_mode_for_record(record: dict[str, Any], target_mode: str) -> str | None:
    if target_mode == SESSION_MODE_ORDER and has_raw_receipt_draft(record):
        return SESSION_MODE_RECEIPT
    if target_mode == SESSION_MODE_RECEIPT and has_raw_order_draft(record):
        return SESSION_MODE_ORDER
    return None


def business_mode_switch_blocked_reply(conflict_mode: str, target_mode: str) -> str:
    return (
        f"现在还有一份{draft_mode_display_name(conflict_mode)}待确认，"
        f"我先不切到{mode_display_name(target_mode)}，避免把数据记串。"
        "请先回复“确认”保存，或发“取消/退出”清掉后再切换。"
    )


def switch_session_mode(user_id: str, mode: str, *, clear_drafts: bool = False) -> None:
    if mode not in SESSION_MODES:
        raise ValueError(f"Invalid session mode: {mode}")

    record = get_session_record(user_id)
    record["mode"] = mode
    if clear_drafts or mode == SESSION_MODE_CHAT:
        record.pop("order_draft", None)
        record.pop("receipt_draft", None)
    elif mode == SESSION_MODE_ORDER:
        record.pop("receipt_draft", None)
    elif mode == SESSION_MODE_RECEIPT:
        record.pop("order_draft", None)
    save_session_record(user_id, record)


def try_switch_business_mode(user_id: str, target_mode: str) -> str | None:
    record = get_session_record(user_id)
    conflict_mode = conflicting_draft_mode_for_record(record, target_mode)
    if conflict_mode:
        return business_mode_switch_blocked_reply(conflict_mode, target_mode)
    switch_session_mode(user_id, target_mode)
    return None


def clear_current_business_draft(user_id: str, mode: str | None = None) -> None:
    mode = mode or get_session_mode(user_id)
    record = get_session_record(user_id)
    if mode == SESSION_MODE_ORDER:
        record.pop("order_draft", None)
    elif mode == SESSION_MODE_RECEIPT:
        record.pop("receipt_draft", None)
    else:
        record.pop("order_draft", None)
        record.pop("receipt_draft", None)
        record["mode"] = SESSION_MODE_CHAT
    save_session_record(user_id, record)


def exit_business_mode(user_id: str) -> str:
    previous_mode = get_session_mode(user_id)
    record = get_session_record(user_id)
    record["mode"] = SESSION_MODE_CHAT
    record.pop("order_draft", None)
    record.pop("receipt_draft", None)
    save_session_record(user_id, record)
    if previous_mode == SESSION_MODE_ORDER:
        return "已退出订单模式，有事随时叫我。"
    if previous_mode == SESSION_MODE_RECEIPT:
        return "已退出入库模式，有事随时叫我。"
    return "现在是普通聊天模式。"


def build_status_message(user_id: str) -> str:
    mode = get_session_mode(user_id)
    record = get_session_record(user_id)
    has_order = has_raw_order_draft(record)
    has_receipt = has_raw_receipt_draft(record)
    if has_order and has_receipt:
        draft_text = "同时发现订单和入库草稿，请先发“取消/退出”清掉后重来"
    elif has_order:
        draft_text = "有一份订单草稿待确认"
    elif has_receipt:
        draft_text = "有一份入库草稿待确认"
    else:
        draft_text = "没有待确认草稿"
    return f"你现在在{mode_display_name(mode)}，{draft_text}。"


def build_mode_help_message(user_id: str) -> str:
    current = mode_display_name(get_session_mode(user_id))
    return (
        f"我现在在{current}。\n"
        "可用模式：\n"
        "1. 普通聊天：正常问事、闲聊两句都可以。\n"
        "2. 订单模式：发“订单”进入，支持订单文字、Excel、照片；确认后写订单库。\n"
        "3. 入库模式：发“入库”进入，发产成品照片；确认后写入库库。\n"
        "常用命令：退出、取消、状态、撤回。"
    )


def build_order_storage_query_reply(user_id: str) -> str:
    draft = get_order_draft(user_id)
    if order_draft_has_content(draft):
        return "现在还有一张订单草稿没入库。确认没问题请回“确认 / 对 / ok / yes”；要改就直接发修改内容。"
    return (
        "我这里不编入库结果。订单只有在你回复“确认入库”后才会写进 orders.db。\n"
        "Web 工具同步时按订单的 order_date 拉取；如果工具是空，先确认刚才那单是否已经保存，以及工具选的下单日期是否和订单里的 order_date 一致。"
    )


ORDER_LIKE_KEYWORDS = {
    "加",
    "订",
    "下单",
    "补",
    "追",
    "改",
    "换",
    "门店",
    "箱",
    "件",
    "袋",
    "盒",
    "包",
}


def looks_like_order_message(message: str) -> bool:
    text = message.strip()
    command = normalize_command(text)
    if not text:
        return False
    if strip_order_inline_prefix(text) is not None:
        return True
    if re.search(r"\d+(?:\.\d+)?\s*(箱|件|袋|盒|包|斤|公斤|kg|KG|份|个|瓶|桶|条|只)", text):
        return True
    if re.search(r"(加|订|下单|补|追|改|换).*\d+", text):
        return True
    if command_contains_any(command, ORDER_LIKE_KEYWORDS) and re.search(r"\d", text):
        return True
    return False


def count_keyword_occurrences(text: str, keywords: set[str]) -> int:
    return sum(text.count(keyword) for keyword in keywords if keyword)


def looks_like_receipt_business_message(message: str) -> bool:
    command = normalize_command(message)
    if is_business_query_or_negated(command) or is_question_like_command(command):
        return False
    if "入库" in command and command_contains_any(command, {"产成品", "成品", "车间", "照片", "图片", "清单", "记录", "记一下"}):
        return True
    return command_contains_any(command, {"产成品", "成品"}) and command_contains_any(command, {"照片", "图片", "车间", "入库"})


def call_global_business_route_llm(messages: list[dict[str, str]]) -> str:
    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=messages,
        temperature=0,
    )
    return (response.choices[0].message.content or "").strip()


def classify_global_business_route(message: str) -> BusinessIntent:
    command = normalize_command(message)
    if is_order_storage_query_command(command):
        return BusinessIntent(GLOBAL_ROUTE_ORDER_QUERY, 0.95, "rule", "order storage query")
    if command_contains_any(command, BUSINESS_NEGATION_KEYWORDS):
        return BusinessIntent(GLOBAL_ROUTE_CHAT, 0.75, "rule", "business negation")
    if looks_like_order_message(message):
        return BusinessIntent(GLOBAL_ROUTE_ORDER_TEXT, 0.92, "rule", "order-like text")
    if looks_like_receipt_business_message(message):
        return BusinessIntent(GLOBAL_ROUTE_ENTER_RECEIPT, 0.9, "rule", "receipt-like text")
    if is_order_mode_command(command):
        return BusinessIntent(GLOBAL_ROUTE_ENTER_ORDER, 0.92, "rule", "order mode command")
    if is_receipt_mode_command(command):
        return BusinessIntent(GLOBAL_ROUTE_ENTER_RECEIPT, 0.92, "rule", "receipt mode command")
    # 规则拿不准 → 交给 agent_router 大脑做大模型分诊。
    # 注意：这里【不再有关键词闸门 should_call_global_business_route_llm】，
    # 所以自然语言 / 复杂订单消息也会进大模型分诊，而不是被默认当成聊天丢掉——
    # 这正是修复"复杂消息看不懂"的关键改动。置信度阈值已内置在 agent_router 中（0.78 / 0.85）。
    decision = agent_router.decide_from_llm(
        message,
        llm_classifier=call_global_business_route_llm,
    )
    return BusinessIntent(decision.intent, decision.confidence, decision.source, decision.reason)


def append_mode_hint(answer: str, mode: str) -> str:
    answer = answer.strip()
    if mode == SESSION_MODE_ORDER:
        return f"{answer}\n\n你还在订单模式，要下单直接发订单文字、Excel 或照片；发“退出”可返回普通聊天。"
    if mode == SESSION_MODE_RECEIPT:
        return f"{answer}\n\n你还在入库模式，要记入库就发产成品照片；发“退出”可返回普通聊天。"
    return answer


def strip_order_inline_prefix(message: str) -> str | None:
    match = re.match(r"^\s*(订单|录单|下单)(?:\s*[：:]|\s+)(.+)$", message, flags=re.DOTALL)
    if not match:
        return None
    return match.group(2).strip()


def build_download_url(path: str) -> str:
    if PUBLIC_BASE_URL:
        return f"{PUBLIC_BASE_URL}{path}"
    return path


def build_order_export_message() -> str:
    url = build_download_url("/exports/orders.xlsx")
    if EXPORT_TOKEN:
        url = f"{url}?token=导出口令"
        return f"订单表可以导出。管理员打开这个地址并填写导出口令：{url}"
    return f"订单表可以导出：{url}"


def get_session_record(user_id: str) -> dict[str, Any]:
    with SESSION_STATE_LOCK:
        state = load_session_state()
        record = state.get(user_id)
        if isinstance(record, dict):
            return dict(record)
    return {}


def save_session_record(user_id: str, record: dict[str, Any]) -> None:
    record["updated_at"] = now_iso()
    with SESSION_STATE_LOCK:
        state = load_session_state()
        state[user_id] = record
        save_session_state(state)


def get_session_mode(user_id: str) -> str:
    mode = str(get_session_record(user_id).get("mode") or SESSION_MODE_CHAT)
    if mode not in SESSION_MODES:
        return SESSION_MODE_CHAT
    return mode


def set_session_mode(user_id: str, mode: str) -> None:
    switch_session_mode(user_id, mode)


def get_order_draft(user_id: str) -> dict[str, Any]:
    draft = get_session_record(user_id).get("order_draft")
    if isinstance(draft, dict):
        return normalize_order_draft(draft)
    return normalize_order_draft({})


def save_order_draft(user_id: str, draft: dict[str, Any]) -> None:
    record = get_session_record(user_id)
    record["mode"] = SESSION_MODE_ORDER
    record["order_draft"] = normalize_order_draft(draft)
    record.pop("receipt_draft", None)
    save_session_record(user_id, record)


def clear_order_draft(user_id: str, *, next_mode: str = SESSION_MODE_ORDER) -> None:
    if next_mode not in SESSION_MODES:
        raise ValueError(f"Invalid session mode: {next_mode}")
    record = get_session_record(user_id)
    record["mode"] = next_mode
    record.pop("order_draft", None)
    if next_mode == SESSION_MODE_CHAT:
        record.pop("receipt_draft", None)
    save_session_record(user_id, record)


def get_receipt_draft(user_id: str) -> dict[str, Any]:
    draft = get_session_record(user_id).get("receipt_draft")
    if isinstance(draft, dict):
        return normalize_receipt_payload(draft)
    return {}


def save_receipt_draft(user_id: str, draft: dict[str, Any]) -> None:
    record = get_session_record(user_id)
    record["mode"] = SESSION_MODE_RECEIPT
    record["receipt_draft"] = normalize_receipt_payload(draft)
    record.pop("order_draft", None)
    save_session_record(user_id, record)


def clear_receipt_draft(user_id: str, *, next_mode: str = SESSION_MODE_RECEIPT) -> None:
    if next_mode not in SESSION_MODES:
        raise ValueError(f"Invalid session mode: {next_mode}")
    record = get_session_record(user_id)
    record["mode"] = next_mode
    record.pop("receipt_draft", None)
    if next_mode == SESSION_MODE_CHAT:
        record.pop("order_draft", None)
    save_session_record(user_id, record)


def call_business_intent_llm(messages: list[dict[str, str]]) -> str:
    response = client.chat.completions.create(
        model=MODEL_NAME,
        messages=messages,
    )
    return (response.choices[0].message.content or "").strip()


def classify_order_reply_intent(message: str, draft: dict[str, Any]) -> BusinessIntent:
    return classify_business_intent(
        message,
        has_draft=order_draft_has_content(draft),
        mode="order",
        draft_summary=format_order_draft_summary(draft),
        llm_classifier=call_business_intent_llm,
    )


def business_confirm_clarification(*, receipt: bool = False) -> str:
    subject = "这条入库记录" if receipt else "这张订单"
    return f"这句话我先不当成确认或修改，{subject}草稿保持不变。确认无误请回“确认 / 对 / ok / yes”；要修改就直接发修改内容。"


def order_draft_reply(prefix: str, draft: dict[str, Any], missing: list[str]) -> str:
    summary = format_order_draft_summary(draft)
    if missing:
        return prefix + "\n" + summary + "\n" + missing_fields_reply(missing)
    return prefix + "\n" + summary + "\n确认无误请回复“确认 / 对 / ok / yes”；要继续修改就直接发修改内容。"


def receipt_draft_reply(prefix: str, draft: dict[str, Any], missing: list[str]) -> str:
    summary = format_receipt_draft_summary(draft)
    if missing:
        return prefix + "\n" + summary + "\n" + missing_fields_reply(missing, receipt=True)
    return prefix + "\n" + summary + "\n确认无误请回复“确认 / 对 / ok / yes”；要继续修改就直接发修改内容。"


def save_confirmed_order_response(user_id: str, draft: dict[str, Any], history_length: int) -> ChatResponse:
    if not order_draft_has_content(draft):
        return ChatResponse(
            user_id=user_id,
            answer="现在没有待确认的订单草稿。直接发订单文字、Excel 或照片都行。",
            history_length=history_length,
        )
    missing = order_draft_missing_fields(draft)
    if missing:
        return ChatResponse(
            user_id=user_id,
            answer=missing_fields_reply(missing),
            history_length=history_length,
        )

    try:
        order_id, line_count = save_confirmed_order(user_id, draft)
    except Exception as exc:
        logger.warning("order_save_failed user_id=%s error=%s", user_id, exc)
        return ChatResponse(
            user_id=user_id,
            answer="订单保存失败了，请稍后再试，或联系管理员查看后台日志。",
            history_length=history_length,
        )

    clear_order_draft(user_id)
    return ChatResponse(
        user_id=user_id,
        answer=f"已保存订单入库，ID {order_id}，共 {line_count} 行商品。继续发下一张即可。",
        history_length=history_length + line_count,
    )


ORDER_SKILL_FILE = Path(
    os.getenv("ORDER_SKILL_FILE", str(Path(__file__).resolve().parent / "skills" / "order" / "SKILL.md"))
)


def load_order_skill() -> str:
    """读取订单 skill（规则 + 例子）。每次读取，方便你改了 .md 立即生效，不用重启想。"""
    try:
        return ORDER_SKILL_FILE.read_text(encoding="utf-8").strip()
    except OSError as exc:
        logger.warning("order_skill_load_failed file=%s error=%s", ORDER_SKILL_FILE, exc)
        return ""


def llm_order_draft_from_message(existing_draft: dict[str, Any], message: str) -> dict[str, Any] | None:
    """订单整理大脑：读 skill + 当前草稿 + 用户消息 → 输出【完整】更新后草稿。

    新单或加 / 改 / 删 / 换都统一走这里：LLM 负责理解并产出完整草稿，
    代码只做归一化与身份字段保护。失败返回 None，由调用方安全兜底。
    """
    skill = load_order_skill()
    today = datetime.now().date().isoformat()
    current = json.dumps(existing_draft or {}, ensure_ascii=False)
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": skill},
                {
                    "role": "user",
                    "content": (
                        f"今天日期：{today}\n\n"
                        f"当前订单草稿（JSON，空对象 {{}} 表示还没有草稿）：\n{current}\n\n"
                        f"用户最新消息：\n{message}\n\n"
                        "请输出更新后的【完整】订单草稿 JSON：应用用户的加 / 改 / 删 / 换，"
                        "保留所有未改动的商品，只输出一个 JSON 对象。"
                    ),
                },
            ],
            temperature=0,
        )
    except Exception as exc:
        logger.warning("order_skill_llm_failed error=%s", exc)
        return None

    parsed = extract_json_object(response.choices[0].message.content or "")
    if not parsed:
        return None

    # 身份 / 系统字段以现有草稿为准，不让文字更新悄悄改掉。
    if existing_draft:
        for key in ("kind", "source", "order_no", "orderer", "raw_ref", "created_at"):
            if existing_draft.get(key):
                parsed[key] = existing_draft.get(key)

    explicit_order_date = extract_explicit_order_date(message)
    if explicit_order_date:
        parsed["order_date"] = explicit_order_date

    parsed["confirmed"] = False
    parsed["status"] = ORDER_STATUS_NEW
    parsed["created_at"] = parsed.get("created_at") or now_iso()

    normalized = normalize_order_draft(parsed)
    if not normalized or not normalized.get("items"):
        return None
    return normalized


def save_confirmed_order(user_id: str, draft: dict[str, Any]) -> tuple[int, int]:
    started_at = time.perf_counter()
    payload = normalize_order_payload(draft)
    missing = order_draft_missing_fields(payload)
    if missing:
        raise ValueError("order draft missing fields: " + ",".join(missing))

    payload["confirmed"] = True
    payload["status"] = ORDER_STATUS_NEW
    payload["raw_ref"] = payload.get("raw_ref") or user_id
    saved = insert_order_payload(payload)
    line_count = len(saved.get("items") or [])
    logger.info(
        "order_confirm_saved user_id=%s order_id=%s source=%s lines=%s elapsed_ms=%s",
        user_id,
        saved.get("id"),
        payload.get("source"),
        line_count,
        int((time.perf_counter() - started_at) * 1000),
    )
    return int(saved["id"]), line_count


def save_confirmed_receipt(draft: dict[str, Any]) -> tuple[str, int]:
    payload = normalize_receipt_payload(draft)
    missing = receipt_missing_fields(payload)
    if missing:
        raise ValueError("receipt draft missing fields: " + ",".join(missing))
    saved = insert_receipt_payload(payload)
    return str(saved["id"]), len(saved.get("items") or [])


def user_order_count(user_id: str) -> int:
    raw_ref_prefixes = (user_id, f"{user_id}:")
    return sum(
        1
        for record in query_order_payloads()
        if str(record.get("raw_ref") or "").startswith(raw_ref_prefixes)
    )


@app.on_event("startup")
async def log_startup() -> None:
    logger.warning(
        "foodwechatbot_startup build=%s file=%s cwd=%s pid=%s database_backend=%s",
        APP_BUILD_LABEL,
        Path(__file__).resolve(),
        Path.cwd(),
        os.getpid(),
        os.getenv("DATABASE_BACKEND", "sqlite"),
    )


def collect_order_records() -> list[dict[str, str]]:
    records = query_order_payloads()

    normalized_records: list[dict[str, str]] = []
    for record in records:
        items = record.get("items") if isinstance(record.get("items"), list) else []
        if not items:
            items = [{}]
        for line_no, item in enumerate(items, start=1):
            if not isinstance(item, dict):
                item = {}
            normalized_records.append(
                {
                    "id": str(record.get("id") or ""),
                    "kind": str(record.get("kind") or ""),
                    "source": str(record.get("source") or ""),
                    "status": str(record.get("status") or ""),
                    "confirmed": "是" if record.get("confirmed") else "否",
                    "store": str(record.get("store") or ""),
                    "order_no": str(record.get("order_no") or ""),
                    "orderer": str(record.get("orderer") or ""),
                    "order_date": str(record.get("order_date") or ""),
                    "deliver_date": str(record.get("deliver_date") or ""),
                    "change_type": str(record.get("change_type") or ""),
                    "line_no": str(line_no),
                    "code": str(item.get("code") or ""),
                    "name": str(item.get("name") or ""),
                    "spec": str(item.get("spec") or ""),
                    "unit": str(item.get("unit") or ""),
                    "qty": "" if item.get("qty") is None else str(item.get("qty")),
                    "price": "" if item.get("price") is None else str(item.get("price")),
                    "category": str(item.get("category") or ""),
                    "raw_text": str(record.get("raw_text") or ""),
                    "raw_ref": str(record.get("raw_ref") or ""),
                    "created_at": str(record.get("created_at") or ""),
                }
            )
    return normalized_records


def trim_history(history: list[dict[str, str]]) -> list[dict[str, str]]:
    return history[-MAX_HISTORY_MESSAGES:]


def build_llm_messages(history: list[dict[str, str]]) -> list[dict[str, str]]:
    messages = [{"role": "system", "content": SYSTEM_PROMPT}]

    if PROMPT_TURN_CONTEXT:
        user_turns = sum(1 for message in history if message.get("role") == "user")
        messages.append(
            {
                "role": "system",
                "content": (
                    f"运行时信息：当前会话已收到 {user_turns} 个用户回合。"
                    "如果系统提示包含阶段流程，请用这个数字判断访谈节奏。"
                    "如果回合数只是建议，不要把它当作硬性拒答或卡死条件。"
                ),
            }
        )

    return [*messages, *history]


def call_llm(user_id: str, history: list[dict[str, str]]) -> str:
    messages = build_llm_messages(history)

    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=messages,
        )
        answer = response.choices[0].message.content
    except Exception as exc:
        print("========== LLM ERROR ==========")
        print(str(exc))
        traceback.print_exc()
        print("================================")
        logger.warning(
            "llm_failed user_id=%s history_length=%s model=%s error_type=%s",
            user_id,
            len(history),
            MODEL_NAME,
            type(exc).__name__,
        )
        raise HTTPException(
            status_code=500,
            detail={
                "error": "LLM request failed",
                "real_error": str(exc),
            },
        ) from exc

    logger.info(
        "llm_success user_id=%s history_length=%s model=%s",
        user_id,
        len(history),
        MODEL_NAME,
    )
    return answer or ""


def call_customer_chat_llm(user_id: str, history: list[dict[str, str]]) -> str:
    messages = [{"role": "system", "content": CUSTOMER_CHAT_PROMPT}, *history]
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=messages,
        )
    except Exception as exc:
        logger.exception(
            "customer_chat_llm_error user_id=%s model=%s error_type=%s",
            user_id,
            MODEL_NAME,
            type(exc).__name__,
        )
        raise HTTPException(
            status_code=500,
            detail={
                "error": "LLM request failed",
                "real_error": str(exc),
            },
        ) from exc

    return (response.choices[0].message.content or "").strip()


def handle_order_user_message(user_id: str, message: str, raw_ref: str | None = None) -> ChatResponse:
    command = normalize_command(message)
    history_length = user_order_count(user_id)
    existing_draft = get_order_draft(user_id)
    has_existing_draft = order_draft_has_content(existing_draft)

    if command in ORDER_CANCEL_COMMANDS:
        clear_order_draft(user_id, next_mode=SESSION_MODE_CHAT)
        return ChatResponse(
            user_id=user_id,
            answer="已清空当前订单草稿，并回到普通聊天。要继续录单再发“订单”。",
            history_length=history_length,
        )

    if has_existing_draft:
        if is_order_draft_view_command(command):
            summary = format_order_draft_summary(existing_draft)
            return ChatResponse(
                user_id=user_id,
                answer=(
                    "当前订单草稿：\n"
                    + summary
                    + "\n确认无误请回复“确认 / 对 / ok / yes”；要修改就直接发修改内容。"
                ),
                history_length=history_length,
            )

        intent = classify_order_reply_intent(message, existing_draft)
        if intent.intent == INTENT_CANCEL and intent.is_rule:
            clear_order_draft(user_id, next_mode=SESSION_MODE_CHAT)
            return ChatResponse(
                user_id=user_id,
                answer="已清空当前订单草稿，并回到普通聊天。要继续录单再发“订单”。",
                history_length=history_length,
            )
        if intent.intent == INTENT_EXIT and intent.is_rule:
            return ChatResponse(
                user_id=user_id,
                answer=exit_business_mode(user_id),
                history_length=0,
            )
        if intent.intent == INTENT_CONFIRM:
            return save_confirmed_order_response(user_id, existing_draft, history_length)
        if intent.intent == INTENT_REJECT:
            return ChatResponse(
                user_id=user_id,
                answer="这单我先不保存。要修改就直接发修改内容；发“取消”可以清空草稿。",
                history_length=history_length,
            )
        if intent.intent == INTENT_MODIFY:
            # 加 / 改 / 删 / 换统一交给订单 skill 大脑，输出完整更新后草稿（旧项不丢）。
            updated_draft = llm_order_draft_from_message(existing_draft, message)
            if not updated_draft:
                return ChatResponse(
                    user_id=user_id,
                    answer="这条修改我没解析成功，麻烦把要加 / 改 / 删的商品和数量说清楚再发一次。",
                    history_length=history_length,
                )
            updated_draft["raw_ref"] = updated_draft.get("raw_ref") or raw_ref or user_id
            if updated_draft.get("kind") == ORDER_KIND_PATCH:
                updated_draft["raw_text"] = message
            save_order_draft(user_id, updated_draft)
            missing = order_draft_missing_fields(updated_draft)
            answer = order_draft_reply("已按你的修改更新订单草稿：", updated_draft, missing)
            return ChatResponse(
                user_id=user_id,
                answer=answer,
                history_length=history_length,
            )
        if intent.intent in {INTENT_UNCLEAR, INTENT_CHAT} and not looks_like_order_message(message):
            return ChatResponse(
                user_id=user_id,
                answer=business_confirm_clarification(),
                history_length=history_length,
            )

    if is_confirm_command(command):
        return ChatResponse(
            user_id=user_id,
            answer="现在没有待确认的订单草稿。直接发订单文字、Excel 或照片都行。",
            history_length=history_length,
        )

    draft = llm_order_draft_from_message(existing_draft, message)
    if not draft:
        return ChatResponse(
            user_id=user_id,
            answer="这条订单我没有解析成功。请把门店、商品、数量说清楚再发一次，例如：老三家 鸡腿 20件。",
            history_length=history_length,
        )

    draft["raw_ref"] = draft.get("raw_ref") or raw_ref or user_id
    if draft.get("kind") == ORDER_KIND_PATCH:
        draft["raw_text"] = draft.get("raw_text") or message
    save_order_draft(user_id, draft)

    missing = order_draft_missing_fields(draft)
    answer = order_draft_reply("我整理成待确认订单：", draft, missing)

    return ChatResponse(
        user_id=user_id,
        answer=answer,
        history_length=history_length,
    )


def format_receipt_draft_summary(draft: dict[str, Any]) -> str:
    if not draft:
        return "暂无入库草稿"
    lines = [
        f"入库日期：{draft.get('date') or '未填写'}",
    ]
    items = draft.get("items") if isinstance(draft.get("items"), list) else []
    if not items:
        lines.append("成品：未填写")
    else:
        lines.append("成品：")
        for index, item in enumerate(items, start=1):
            if not isinstance(item, dict):
                continue
            parts = [
                str(item.get("code") or "").strip(),
                str(item.get("name") or "未填写成品").strip(),
                str(item.get("spec") or "").strip(),
                f"{item.get('qty') if item.get('qty') is not None else '未填写数量'}{item.get('unit') or ''}",
            ]
            lines.append(f"{index}. {' / '.join(part for part in parts if part)}")
    return "\n".join(lines)


RECEIPT_SKILL_FILE = Path(
    os.getenv("RECEIPT_SKILL_FILE", str(Path(__file__).resolve().parent / "skills" / "receipt" / "SKILL.md"))
)


def load_receipt_skill() -> str:
    """读取入库 skill（规则 + 例子）。每次读取，改了 .md 立即生效。"""
    try:
        return RECEIPT_SKILL_FILE.read_text(encoding="utf-8").strip()
    except OSError as exc:
        logger.warning("receipt_skill_load_failed file=%s error=%s", RECEIPT_SKILL_FILE, exc)
        return ""


def llm_receipt_draft_from_message(existing_draft: dict[str, Any], message: str) -> dict[str, Any] | None:
    """入库整理大脑：读 skill + 当前草稿 + 用户消息 → 输出【完整】更新后入库草稿。

    加 / 改 / 删 / 换统一走这里；代码只做归一化与身份字段保护。失败返回 None，调用方兜底。
    """
    skill = load_receipt_skill()
    today = datetime.now().date().isoformat()
    current = json.dumps(existing_draft or {}, ensure_ascii=False)
    try:
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": skill},
                {
                    "role": "user",
                    "content": (
                        f"今天日期：{today}\n\n"
                        f"当前入库草稿（JSON）：\n{current}\n\n"
                        f"用户最新消息：\n{message}\n\n"
                        "请输出更新后的【完整】入库草稿 JSON：应用用户的加 / 改 / 删 / 换，"
                        "保留所有未改动的成品，只输出一个 JSON 对象。"
                    ),
                },
            ],
            temperature=0,
        )
    except Exception as exc:
        logger.warning("receipt_skill_llm_failed error=%s", exc)
        return None

    parsed = extract_json_object(response.choices[0].message.content or "")
    if not parsed:
        return None

    # 身份 / 系统字段以现有草稿为准，不让文字更新悄悄改掉。
    if existing_draft:
        for key in ("status", "created_at", "id", "raw_ref"):
            if existing_draft.get(key):
                parsed[key] = existing_draft.get(key)

    normalized = normalize_receipt_payload(parsed)
    if not normalized or not normalized.get("items"):
        return None
    return normalized


def classify_receipt_reply_intent(message: str, draft: dict[str, Any]) -> BusinessIntent:
    return classify_business_intent(
        message,
        has_draft=receipt_draft_has_content(draft),
        mode="receipt",
        draft_summary=format_receipt_draft_summary(draft),
        llm_classifier=call_business_intent_llm,
    )


def save_confirmed_receipt_response(user_id: str, draft: dict[str, Any]) -> ChatResponse:
    if not receipt_draft_has_content(draft):
        return ChatResponse(
            user_id=user_id,
            answer="现在没有待确认的入库草稿。发产成品入库照片给我就行。",
            history_length=0,
        )
    missing = receipt_missing_fields(draft)
    if missing:
        return ChatResponse(
            user_id=user_id,
            answer=missing_fields_reply(missing, receipt=True),
            history_length=0,
        )

    try:
        receipt_id, line_count = save_confirmed_receipt(draft)
    except Exception as exc:
        logger.warning("receipt_save_failed user_id=%s error=%s", user_id, exc)
        return ChatResponse(
            user_id=user_id,
            answer="入库记录保存失败了，请稍后再试，或联系管理员查看后台日志。",
            history_length=0,
        )

    clear_receipt_draft(user_id)
    return ChatResponse(
        user_id=user_id,
        answer=f"已保存产成品入库记录 {receipt_id}，共 {line_count} 行成品。",
        history_length=0,
    )


def handle_receipt_user_message(user_id: str, message: str) -> ChatResponse:
    command = normalize_command(message)
    draft = get_receipt_draft(user_id)
    has_existing_draft = receipt_draft_has_content(draft)
    if command in ORDER_CANCEL_COMMANDS:
        clear_receipt_draft(user_id, next_mode=SESSION_MODE_CHAT)
        return ChatResponse(
            user_id=user_id,
            answer="已清空当前入库草稿，并回到普通聊天。要继续入库再发“入库”。",
            history_length=0,
        )

    if has_existing_draft:
        intent = classify_receipt_reply_intent(message, draft)
        if intent.intent == INTENT_CANCEL and intent.is_rule:
            clear_receipt_draft(user_id, next_mode=SESSION_MODE_CHAT)
            return ChatResponse(
                user_id=user_id,
                answer="已清空当前入库草稿，并回到普通聊天。要继续入库再发“入库”。",
                history_length=0,
            )
        if intent.intent == INTENT_EXIT and intent.is_rule:
            return ChatResponse(
                user_id=user_id,
                answer=exit_business_mode(user_id),
                history_length=0,
            )
        if intent.intent == INTENT_CONFIRM:
            return save_confirmed_receipt_response(user_id, draft)
        if intent.intent == INTENT_REJECT:
            return ChatResponse(
                user_id=user_id,
                answer="这条入库记录我先不保存。要修改请重新发送照片；发“取消”可以清空草稿。",
                history_length=0,
            )
        if intent.intent == INTENT_MODIFY:
            # 加 / 改 / 删 / 换统一交给入库 skill 大脑，输出完整更新后草稿（旧成品不丢）。
            updated_draft = llm_receipt_draft_from_message(draft, message)
            if not updated_draft:
                return ChatResponse(
                    user_id=user_id,
                    answer="这条修改我没解析成功，麻烦把要加 / 改 / 删的成品和数量说清楚再发一次。",
                    history_length=0,
                )
            save_receipt_draft(user_id, updated_draft)
            missing = receipt_missing_fields(updated_draft)
            answer = receipt_draft_reply("已按你的修改更新入库草稿：", updated_draft, missing)
            return ChatResponse(user_id=user_id, answer=answer, history_length=0)
        if intent.intent in {INTENT_UNCLEAR, INTENT_CHAT}:
            return ChatResponse(
                user_id=user_id,
                answer=business_confirm_clarification(receipt=True),
                history_length=0,
            )

    if is_confirm_command(command):
        return ChatResponse(
            user_id=user_id,
            answer="现在没有待确认的入库草稿。发产成品入库照片给我就行。",
            history_length=0,
        )

    return ChatResponse(
        user_id=user_id,
        answer="当前是入库模式。请发送产成品入库照片；识别后我会发清单给你确认。发“订单”可切到订单模式。",
        history_length=0,
    )


def handle_receipt_photo_input(user_id: str, image_bytes: bytes, mime_type: str | None, raw_ref: str) -> ChatResponse:
    try:
        draft = llm_parse_receipt_photo(vision_client, VISION_MODEL, image_bytes, mime_type, raw_ref)
    except Exception as exc:
        logger.warning("receipt_photo_parse_failed user_id=%s raw_ref=%s error=%s", user_id, raw_ref, exc)
        if "vision model is not configured" in str(exc):
            answer = "入库照片已收到，但当前视觉模型还没配置好。请稍后再试，或先人工记录。"
        else:
            answer = "这张入库照片我没有识别成功。请重新拍清楚成品名称和数量后再发。"
        return ChatResponse(user_id=user_id, answer=answer, history_length=0)

    save_receipt_draft(user_id, draft)
    missing = receipt_missing_fields(draft)
    answer = receipt_draft_reply("我把照片识别成待确认入库记录：", draft, missing)
    return ChatResponse(user_id=user_id, answer=answer, history_length=0)


def needs_human_transfer(message: str) -> bool:
    command = normalize_command(message)
    return command_contains_any(
        command,
        {"价格", "多少钱", "报价", "投诉", "赔", "纠纷", "发票", "退款", "催单", "交期", "合同"},
    )


def handle_general_chat(user_id: str, message: str, mode_hint: str | None = None) -> ChatResponse:
    if needs_human_transfer(message):
        answer = "这个我不瞎承诺，我帮您转人工处理。"
        if mode_hint:
            answer = append_mode_hint(answer, mode_hint)
        return ChatResponse(user_id=user_id, answer=answer, history_length=0)

    memory_key = f"customer_chat:{user_id}"
    with MEMORY_LOCK:
        memory = load_memory()
        history = memory.setdefault(memory_key, [])

        if not isinstance(history, list):
            raise HTTPException(status_code=500, detail=f"Invalid history for user_id: {user_id}")

        user_message_ts = time.time()
        history.append(
            {
                "role": "user",
                "content": message,
                "created_at": now_iso(),
                "ts": user_message_ts,
            }
        )
        history = trim_history(history)
        memory[memory_key] = history

        logger.info("customer_chat_request user_id=%s history_length=%s", user_id, len(history))

        answer = call_customer_chat_llm(user_id, history).strip()
        if mode_hint:
            answer = append_mode_hint(answer, mode_hint)

        history.append(
            {
                "role": "assistant",
                "content": answer,
                "created_at": now_iso(),
                "ts": time.time(),
            }
        )
        history = trim_history(history)
        memory[memory_key] = history
        save_memory(memory)

    return ChatResponse(
        user_id=user_id,
        answer=answer,
        history_length=len(history),
    )


def handle_user_message(user_id: str, message: str, raw_ref: str | None = None) -> ChatResponse:
    user_id = user_id.strip()
    message = message.strip()

    if not user_id:
        raise HTTPException(status_code=400, detail="user_id cannot be empty")
    if not message:
        raise HTTPException(status_code=400, detail="message cannot be empty")

    command = normalize_command(message)
    inline_order_message = strip_order_inline_prefix(message)
    current_mode = get_session_mode(user_id)
    record = get_session_record(user_id)
    has_order_draft = has_raw_order_draft(record)
    has_receipt_draft = has_raw_receipt_draft(record)
    has_any_draft = has_order_draft or has_receipt_draft

    if command in ORDER_EXPORT_COMMANDS:
        return ChatResponse(
            user_id=user_id,
            answer=build_order_export_message(),
            history_length=user_order_count(user_id),
        )

    if is_status_command(command):
        return ChatResponse(
            user_id=user_id,
            answer=build_status_message(user_id),
            history_length=user_order_count(user_id),
        )

    if is_mode_help_command(command):
        return ChatResponse(
            user_id=user_id,
            answer=build_mode_help_message(user_id),
            history_length=user_order_count(user_id),
        )

    if is_exit_mode_command(command):
        return ChatResponse(
            user_id=user_id,
            answer=exit_business_mode(user_id),
            history_length=0,
        )

    if command in ORDER_CANCEL_COMMANDS:
        if has_receipt_draft and not has_order_draft:
            clear_receipt_draft(user_id, next_mode=SESSION_MODE_CHAT)
            answer = "已清空当前入库草稿，并回到普通聊天。"
        elif has_order_draft and not has_receipt_draft:
            clear_order_draft(user_id, next_mode=SESSION_MODE_CHAT)
            answer = "已清空当前订单草稿，并回到普通聊天。"
        else:
            record = get_session_record(user_id)
            record.pop("order_draft", None)
            record.pop("receipt_draft", None)
            record["mode"] = SESSION_MODE_CHAT
            save_session_record(user_id, record)
            answer = "已清空当前草稿，并回到普通聊天。"
        return ChatResponse(user_id=user_id, answer=answer, history_length=0)

    if is_revoke_command(command):
        if current_mode == SESSION_MODE_RECEIPT or is_receipt_revoke_target(command):
            answer = cancel_latest_receipt_for_user(user_id)
        else:
            answer = cancel_latest_order_for_user(user_id)
        return ChatResponse(user_id=user_id, answer=answer, history_length=user_order_count(user_id))

    if needs_human_transfer(message):
        return ChatResponse(user_id=user_id, answer="这个我不瞎承诺，我帮您转人工处理。", history_length=0)

    if has_order_draft and has_receipt_draft:
        return ChatResponse(
            user_id=user_id,
            answer="我发现当前同时有订单草稿和入库草稿，状态不安全。请发“取消”清空后重新录入，避免把数据记串。",
            history_length=0,
        )

    if is_order_mode_command(command) and has_receipt_draft:
        return ChatResponse(
            user_id=user_id,
            answer=business_mode_switch_blocked_reply(SESSION_MODE_RECEIPT, SESSION_MODE_ORDER),
            history_length=0,
        )
    if is_receipt_mode_command(command) and has_order_draft:
        return ChatResponse(
            user_id=user_id,
            answer=business_mode_switch_blocked_reply(SESSION_MODE_ORDER, SESSION_MODE_RECEIPT),
            history_length=user_order_count(user_id),
        )

    if has_order_draft:
        return handle_order_user_message(user_id, message, raw_ref=raw_ref)
    if has_receipt_draft:
        return handle_receipt_user_message(user_id, message)

    if is_confirm_command(command, has_draft=has_any_draft):
        if current_mode == SESSION_MODE_ORDER:
            return handle_order_user_message(user_id, message, raw_ref=raw_ref)
        if current_mode == SESSION_MODE_RECEIPT:
            return handle_receipt_user_message(user_id, message)
        return ChatResponse(
            user_id=user_id,
            answer="现在没有待确认的业务草稿。要录订单发“订单”，要记入库发“入库”。",
            history_length=0,
        )

    if is_order_mode_command(command):
        blocked = try_switch_business_mode(user_id, SESSION_MODE_ORDER)
        if blocked:
            return ChatResponse(user_id=user_id, answer=blocked, history_length=user_order_count(user_id))
        return ChatResponse(
            user_id=user_id,
            answer="好的，进入订单模式了，直接发订单文字、Excel 或照片都行，发“退出”就退出。",
            history_length=user_order_count(user_id),
        )

    if is_receipt_mode_command(command):
        blocked = try_switch_business_mode(user_id, SESSION_MODE_RECEIPT)
        if blocked:
            return ChatResponse(user_id=user_id, answer=blocked, history_length=0)
        return ChatResponse(
            user_id=user_id,
            answer="好的，进入入库模式了，发产成品入库照片就行，发“退出”就退出。",
            history_length=0,
        )

    if is_order_storage_query_command(command):
        return ChatResponse(
            user_id=user_id,
            answer=build_order_storage_query_reply(user_id),
            history_length=user_order_count(user_id),
        )

    if inline_order_message is not None:
        blocked = try_switch_business_mode(user_id, SESSION_MODE_ORDER)
        if blocked:
            return ChatResponse(user_id=user_id, answer=blocked, history_length=user_order_count(user_id))
        return handle_order_user_message(user_id, inline_order_message, raw_ref=raw_ref)

    if current_mode == SESSION_MODE_ORDER:
        # 订单模式里也不靠关键词硬判：一律问分诊大脑（规则在 skills/routing/SKILL.md，你可改）。
        # 只有判为"给了具体商品内容"才录单；问句/想改但没给内容 → 自然聊天接住，不怼"解析失败"。
        order_mode_route = agent_router.decide_from_llm(
            message,
            mode=SESSION_MODE_ORDER,
            llm_classifier=call_global_business_route_llm,
        )
        if order_mode_route.intent == agent_router.ROUTE_ORDER_TEXT:
            return handle_order_user_message(user_id, message, raw_ref=raw_ref)
        return handle_general_chat(user_id, message, mode_hint=SESSION_MODE_ORDER)

    if current_mode == SESSION_MODE_RECEIPT:
        return handle_general_chat(user_id, message, mode_hint=SESSION_MODE_RECEIPT)

    route = classify_global_business_route(message)
    if route.intent == GLOBAL_ROUTE_ORDER_TEXT:
        blocked = try_switch_business_mode(user_id, SESSION_MODE_ORDER)
        if blocked:
            return ChatResponse(user_id=user_id, answer=blocked, history_length=user_order_count(user_id))
        return handle_order_user_message(user_id, message, raw_ref=raw_ref)
    if route.intent == GLOBAL_ROUTE_ENTER_ORDER:
        blocked = try_switch_business_mode(user_id, SESSION_MODE_ORDER)
        if blocked:
            return ChatResponse(user_id=user_id, answer=blocked, history_length=user_order_count(user_id))
        return ChatResponse(
            user_id=user_id,
            answer="好的，进入订单模式了，直接发订单文字、Excel 或照片都行，发“退出”就退出。",
            history_length=user_order_count(user_id),
        )
    if route.intent == GLOBAL_ROUTE_ENTER_RECEIPT:
        blocked = try_switch_business_mode(user_id, SESSION_MODE_RECEIPT)
        if blocked:
            return ChatResponse(user_id=user_id, answer=blocked, history_length=0)
        return ChatResponse(
            user_id=user_id,
            answer="好的，进入入库模式了，发产成品入库照片就行，发“退出”就退出。",
            history_length=0,
        )
    if route.intent == GLOBAL_ROUTE_ORDER_QUERY:
        return ChatResponse(
            user_id=user_id,
            answer=build_order_storage_query_reply(user_id),
            history_length=user_order_count(user_id),
        )

    return handle_general_chat(user_id, message)


def get_wecom_crypto() -> WXBizMsgCrypt:
    if not WECOM_CALLBACK_TOKEN or not WECOM_ENCODING_AES_KEY:
        raise HTTPException(
            status_code=500,
            detail="WECOM_CALLBACK_TOKEN and WECOM_ENCODING_AES_KEY must be configured",
        )

    return WXBizMsgCrypt(
        WECOM_CALLBACK_TOKEN,
        WECOM_ENCODING_AES_KEY,
        WECOM_CORP_ID,
        channel=WxChannel_Wecom,
    )


def require_query_param(request: Request, name: str) -> str:
    value = request.query_params.get(name)
    if not value:
        raise HTTPException(status_code=400, detail=f"missing query parameter: {name}")
    return value


def xml_text(parent: ET.Element, path: str, default: str = "") -> str:
    node = parent.find(path)
    if node is None or node.text is None:
        return default
    return node.text


def parse_wecom_plain_message(plain_xml: bytes | str) -> WecomMessage:
    if isinstance(plain_xml, bytes):
        plain_xml = plain_xml.decode("utf-8")

    root = ET.fromstring(plain_xml)
    msg_type = xml_text(root, "MsgType")
    chat_type = xml_text(root, "ChatType")
    chat_id = xml_text(root, "ChatId")
    msg_id = xml_text(root, "MsgId")
    sender_user_id = xml_text(root, "From/UserId")
    sender_name = xml_text(root, "From/Name") or xml_text(root, "From/Alias") or sender_user_id

    if msg_type == "text":
        content = xml_text(root, "Text/Content")
    elif msg_type == "mixed":
        parts: list[str] = []
        mixed = root.find("MixedMessage")
        if mixed is not None:
            for item in mixed:
                if xml_text(item, "MsgType") == "text":
                    parts.append(xml_text(item, "Text/Content"))
        content = "\n".join(part for part in parts if part).strip()
    elif msg_type == "event":
        event_type = xml_text(root, "Event/EventType")
        content = f"[event:{event_type}]"
    else:
        content = ""

    return WecomMessage(
        msg_type=msg_type,
        chat_type=chat_type,
        chat_id=chat_id,
        msg_id=msg_id,
        sender_user_id=sender_user_id,
        sender_name=sender_name,
        content=content,
    )


def strip_bot_mention(content: str) -> str:
    content = content.strip()
    if WECOM_BOT_NAME:
        content = content.replace(f"@{WECOM_BOT_NAME}", "").strip()
    return content


def build_wecom_text_response_xml(content: str) -> bytes:
    root = ET.Element("xml")
    ET.SubElement(root, "MsgType").text = "text"
    text = ET.SubElement(root, "Text")
    ET.SubElement(text, "Content").text = content
    return ET.tostring(root, encoding="utf-8", method="xml")


def encrypt_wecom_response(content: str, nonce: str, timestamp: str) -> str:
    ret, encrypted_response = get_wecom_crypto().EncryptMsg(
        build_wecom_text_response_xml(content),
        nonce,
        timestamp,
    )
    if ret != 0:
        logger.warning("wecom_encrypt_failed ret=%s", ret)
        raise HTTPException(status_code=500, detail=f"WeCom encrypt failed: {ret}")

    if isinstance(encrypted_response, bytes):
        return encrypted_response.decode("utf-8")
    return encrypted_response


def is_duplicate_wecom_message(msg_id: str) -> bool:
    if not msg_id:
        return False

    with SEEN_WECOM_MSG_IDS_LOCK:
        if msg_id in SEEN_WECOM_MSG_IDS:
            return True
        SEEN_WECOM_MSG_IDS.add(msg_id)
        if len(SEEN_WECOM_MSG_IDS) > 1000:
            for old_msg_id in list(SEEN_WECOM_MSG_IDS)[:200]:
                SEEN_WECOM_MSG_IDS.discard(old_msg_id)
        return False


def answer_wecom_message(message: WecomMessage) -> str:
    if message.msg_type == "event":
        return "我已加入，可以在群里 @我 提问。"

    if message.msg_type not in {"text", "mixed"}:
        return "目前我先支持文字消息，图片、文件后面再接。"

    content = strip_bot_mention(message.content)
    if not content:
        return "我在，直接说你的问题。"

    if message.chat_type == "group":
        memory_user_id = f"group:{message.chat_id}"
        llm_message = f"{message.sender_name}: {content}"
    else:
        memory_user_id = f"user:{message.sender_user_id or message.chat_id}"
        llm_message = content

    raw_ref = f"wecom:{message.chat_id or message.sender_user_id}:{message.msg_id}"
    return handle_user_message(memory_user_id, llm_message, raw_ref=raw_ref).answer


def get_wecom_kf_crypto() -> WXBizMsgCrypt:
    if not WECOM_KF_CORP_ID:
        raise HTTPException(status_code=500, detail="WECOM_KF_CORP_ID must be configured")
    if not WECOM_KF_CALLBACK_TOKEN or not WECOM_KF_ENCODING_AES_KEY:
        raise HTTPException(
            status_code=500,
            detail="WECOM_KF_CALLBACK_TOKEN and WECOM_KF_ENCODING_AES_KEY must be configured",
        )

    return WXBizMsgCrypt(
        WECOM_KF_CALLBACK_TOKEN,
        WECOM_KF_ENCODING_AES_KEY,
        WECOM_KF_CORP_ID,
        channel=WxChannel_Wecom,
    )


def compute_wecom_signature(token: str, timestamp: str, nonce: str, encrypted: str) -> str:
    raw = "".join(sorted([token, timestamp, nonce, encrypted]))
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def decrypt_wecom_kf_payload(encrypted: str) -> tuple[bytes, str]:
    if not WECOM_KF_ENCODING_AES_KEY:
        raise HTTPException(status_code=500, detail="WECOM_KF_ENCODING_AES_KEY must be configured")

    try:
        key = base64.b64decode(WECOM_KF_ENCODING_AES_KEY + "=")
        if len(key) != 32:
            raise ValueError("invalid key length")
        cryptor = AES.new(key, AES.MODE_CBC, key[:16])
        plain_text = cryptor.decrypt(base64.b64decode(encrypted))
        pad = plain_text[-1]
        if pad < 1 or pad > 32:
            raise ValueError("invalid padding")

        content = plain_text[16:-pad]
        if len(content) < 4:
            raise ValueError("missing message length")
        xml_len = socket.ntohl(struct.unpack("I", content[:4])[0])
        xml_content = content[4 : xml_len + 4]
        receive_id = content[xml_len + 4 :].decode("utf-8", errors="replace")
        return xml_content, receive_id
    except Exception as exc:
        logger.warning("wecom_kf_decrypt_payload_failed error=%s", exc)
        raise HTTPException(status_code=403, detail="WeCom KF decrypt failed") from exc


def verify_wecom_kf_signature(msg_signature: str, timestamp: str, nonce: str, encrypted: str) -> None:
    if not WECOM_KF_CALLBACK_TOKEN:
        raise HTTPException(status_code=500, detail="WECOM_KF_CALLBACK_TOKEN must be configured")

    expected_signature = compute_wecom_signature(
        WECOM_KF_CALLBACK_TOKEN,
        timestamp,
        nonce,
        encrypted,
    )
    if not hmac.compare_digest(expected_signature, msg_signature):
        logger.warning(
            "wecom_kf_signature_failed expected=%s actual=%s",
            expected_signature,
            msg_signature,
        )
        raise HTTPException(status_code=403, detail="WeCom KF signature failed")


def decrypt_wecom_kf_verify_url(
    msg_signature: str,
    timestamp: str,
    nonce: str,
    echostr: str,
) -> str:
    verify_wecom_kf_signature(msg_signature, timestamp, nonce, echostr)
    decrypted_echo, receive_id = decrypt_wecom_kf_payload(echostr)
    if WECOM_KF_CORP_ID and receive_id != WECOM_KF_CORP_ID:
        logger.warning(
            "wecom_kf_receive_id_mismatch expected=%s actual=%s",
            WECOM_KF_CORP_ID,
            receive_id,
        )
    return decrypted_echo.decode("utf-8")


def decrypt_wecom_kf_message(
    encrypted_body: bytes,
    msg_signature: str,
    timestamp: str,
    nonce: str,
) -> bytes:
    try:
        encrypted = xml_text(ET.fromstring(encrypted_body), "Encrypt")
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Invalid WeCom KF encrypted XML") from exc

    if not encrypted:
        raise HTTPException(status_code=400, detail="Missing WeCom KF Encrypt field")

    verify_wecom_kf_signature(msg_signature, timestamp, nonce, encrypted)
    plain_xml, receive_id = decrypt_wecom_kf_payload(encrypted)
    if WECOM_KF_CORP_ID and receive_id != WECOM_KF_CORP_ID:
        logger.warning(
            "wecom_kf_receive_id_mismatch expected=%s actual=%s",
            WECOM_KF_CORP_ID,
            receive_id,
        )
    return plain_xml


def parse_wecom_kf_event(plain_xml: bytes | str) -> WecomKfEvent:
    if isinstance(plain_xml, bytes):
        plain_xml = plain_xml.decode("utf-8")

    root = ET.fromstring(plain_xml)
    msg_type = xml_text(root, "MsgType")
    event = xml_text(root, "Event")

    if msg_type != "event" or event != "kf_msg_or_event":
        raise ValueError(f"Unsupported WeCom KF callback: MsgType={msg_type}, Event={event}")

    return WecomKfEvent(
        token=xml_text(root, "Token"),
        open_kfid=xml_text(root, "OpenKfId"),
        event=event,
        create_time=xml_text(root, "CreateTime"),
    )


def get_wecom_kf_access_token() -> str:
    global WECOM_KF_ACCESS_TOKEN, WECOM_KF_ACCESS_TOKEN_EXPIRES_AT

    now = time.time()
    with WECOM_KF_ACCESS_TOKEN_LOCK:
        if WECOM_KF_ACCESS_TOKEN and WECOM_KF_ACCESS_TOKEN_EXPIRES_AT > now + 60:
            return WECOM_KF_ACCESS_TOKEN

        if not WECOM_KF_CORP_ID or not WECOM_KF_SECRET:
            raise RuntimeError("WECOM_KF_CORP_ID and WECOM_KF_SECRET must be configured")

        response = httpx.get(
            "https://qyapi.weixin.qq.com/cgi-bin/gettoken",
            params={
                "corpid": WECOM_KF_CORP_ID,
                "corpsecret": WECOM_KF_SECRET,
            },
            timeout=HTTP_TIMEOUT_SECONDS,
        )
        response.raise_for_status()
        data = response.json()

        if data.get("errcode") != 0:
            raise RuntimeError(f"WeCom gettoken failed: {data}")

        access_token = data.get("access_token")
        if not access_token:
            raise RuntimeError(f"WeCom gettoken missing access_token: {data}")

        WECOM_KF_ACCESS_TOKEN = access_token
        WECOM_KF_ACCESS_TOKEN_EXPIRES_AT = now + int(data.get("expires_in", 7200)) - 300
        return WECOM_KF_ACCESS_TOKEN


def load_kf_cursors() -> dict[str, str]:
    if models.is_enabled():
        return models.load_kf_cursors()

    if not WECOM_KF_CURSOR_FILE.exists():
        return {}

    raw_cursors = WECOM_KF_CURSOR_FILE.read_text(encoding="utf-8").strip()
    if not raw_cursors:
        return {}

    data = json.loads(raw_cursors)
    if not isinstance(data, dict):
        logger.warning("kf_cursor_file_invalid path=%s", WECOM_KF_CURSOR_FILE)
        return {}

    return {str(key): str(value) for key, value in data.items() if value is not None}


def save_kf_cursors(cursors: dict[str, str]) -> None:
    if models.is_enabled():
        models.save_kf_cursors(cursors)
        return

    WECOM_KF_CURSOR_FILE.write_text(
        json.dumps(cursors, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def get_kf_cursor(open_kfid: str) -> str:
    if not open_kfid:
        return ""

    with WECOM_KF_CURSOR_LOCK:
        return load_kf_cursors().get(open_kfid, "")


def set_kf_cursor(open_kfid: str, cursor: str) -> None:
    if not open_kfid or not cursor:
        return

    with WECOM_KF_CURSOR_LOCK:
        cursors = load_kf_cursors()
        cursors[open_kfid] = cursor
        save_kf_cursors(cursors)


def is_duplicate_wecom_kf_message(msg_id: str) -> bool:
    if not msg_id:
        return False

    with SEEN_WECOM_KF_MSG_IDS_LOCK:
        if msg_id in SEEN_WECOM_KF_MSG_IDS:
            return True
        SEEN_WECOM_KF_MSG_IDS.add(msg_id)
        if len(SEEN_WECOM_KF_MSG_IDS) > 5000:
            for old_msg_id in list(SEEN_WECOM_KF_MSG_IDS)[:1000]:
                SEEN_WECOM_KF_MSG_IDS.discard(old_msg_id)
        return False


def post_wecom_kf_api(path: str, payload: dict[str, Any]) -> dict[str, Any]:
    access_token = get_wecom_kf_access_token()
    response = httpx.post(
        f"https://qyapi.weixin.qq.com/cgi-bin/{path}",
        params={"access_token": access_token},
        json=payload,
        timeout=HTTP_TIMEOUT_SECONDS,
    )
    response.raise_for_status()
    data = response.json()
    if data.get("errcode") != 0:
        raise WecomKfApiError(path, data)
    return data


def get_wecom_kf_service_state(open_kfid: str, external_userid: str) -> int | None:
    data = post_wecom_kf_api(
        "kf/service_state/get",
        {
            "open_kfid": open_kfid,
            "external_userid": external_userid,
        },
    )
    service_state = data.get("service_state")
    try:
        state = int(service_state)
    except (TypeError, ValueError):
        logger.warning(
            "wecom_kf_service_state_invalid open_kfid=%s external_userid=%s data=%s",
            open_kfid,
            external_userid,
            data,
        )
        return None

    logger.info(
        "wecom_kf_service_state open_kfid=%s external_userid=%s state=%s servicer=%s",
        open_kfid,
        external_userid,
        state,
        data.get("servicer_userid"),
    )
    return state


def transfer_wecom_kf_to_ai(open_kfid: str, external_userid: str) -> None:
    data = post_wecom_kf_api(
        "kf/service_state/trans",
        {
            "open_kfid": open_kfid,
            "external_userid": external_userid,
            "service_state": 1,
        },
    )
    logger.info(
        "wecom_kf_service_state_trans_to_ai open_kfid=%s external_userid=%s msg_code=%s",
        open_kfid,
        external_userid,
        data.get("msg_code"),
    )


def ensure_wecom_kf_ai_session(open_kfid: str, external_userid: str) -> None:
    state = get_wecom_kf_service_state(open_kfid, external_userid)
    if state in (0, 1, None):
        return

    try:
        transfer_wecom_kf_to_ai(open_kfid, external_userid)
    except WecomKfApiError as exc:
        logger.warning(
            "wecom_kf_service_state_trans_to_ai_failed open_kfid=%s external_userid=%s state=%s errcode=%s data=%s",
            open_kfid,
            external_userid,
            state,
            exc.errcode,
            exc.data,
        )
        raise


def send_wecom_kf_text(open_kfid: str, external_userid: str, content: str) -> None:
    if not open_kfid or not external_userid:
        raise RuntimeError("open_kfid and external_userid are required to send WeCom KF message")

    payload = {
        "touser": external_userid,
        "open_kfid": open_kfid,
        "msgtype": "text",
        "text": {"content": content},
    }
    ensure_wecom_kf_ai_session(open_kfid, external_userid)

    try:
        post_wecom_kf_api("kf/send_msg", payload)
    except WecomKfApiError as exc:
        if exc.errcode != 95018:
            raise

        logger.warning(
            "wecom_kf_send_state_invalid_retry open_kfid=%s external_userid=%s data=%s",
            open_kfid,
            external_userid,
            exc.data,
        )
        transfer_wecom_kf_to_ai(open_kfid, external_userid)
        post_wecom_kf_api("kf/send_msg", payload)

    logger.info("wecom_kf_send_success open_kfid=%s external_userid=%s", open_kfid, external_userid)


def get_wecom_kf_media(media_id: str) -> tuple[bytes, str, str]:
    if not media_id:
        raise RuntimeError("media_id is required")

    access_token = get_wecom_kf_access_token()
    response = httpx.get(
        "https://qyapi.weixin.qq.com/cgi-bin/media/get",
        params={"access_token": access_token, "media_id": media_id},
        timeout=HTTP_TIMEOUT_SECONDS,
        follow_redirects=True,
    )
    response.raise_for_status()

    content_type = response.headers.get("content-type", "").split(";")[0].strip()
    if content_type == "application/json":
        try:
            data = response.json()
        except ValueError as exc:
            raise RuntimeError("WeCom media/get returned invalid JSON") from exc
        raise RuntimeError(f"WeCom media/get failed: {data}")

    filename = ""
    disposition = response.headers.get("content-disposition", "")
    encoded_match = re.search(r"filename\*=(?:UTF-8''|utf-8'')?([^;]+)", disposition)
    if encoded_match:
        filename = unquote(encoded_match.group(1).strip().strip('"'))
    match = re.search(r'filename="?([^";]+)"?', disposition)
    if match and not filename:
        filename = match.group(1)
    return response.content, content_type, filename


def save_excel_order_payloads(file_bytes: bytes, raw_ref: str) -> list[dict[str, Any]]:
    payloads = parse_excel_order_payloads(file_bytes, raw_ref)
    saved: list[dict[str, Any]] = []
    for payload in payloads:
        payload["confirmed"] = True
        payload["status"] = ORDER_STATUS_NEW
        saved.append(insert_order_payload(payload))
    return saved


def format_saved_order_ids(saved_orders: list[dict[str, Any]]) -> str:
    ids = [str(order.get("id")) for order in saved_orders if order.get("id")]
    return "、".join(ids)


def excel_order_failure_reply(exc: Exception) -> str:
    error_text = str(exc)
    if any(marker in error_text for marker in ("not Excel", "not a valid .xlsx", "File is not a zip file")):
        return "Excel订单解析失败了。我这边收到的文件内容不像有效的 .xlsx，请重新发送原始 Excel 文件。"
    if "legacy .xls" in error_text:
        return "Excel订单解析失败了。当前只支持 .xlsx/.xlsm，请先把表格另存为 .xlsx 后再发。"
    if any(marker in error_text for marker in ("no order item rows", "header row not found")):
        return "Excel订单解析失败了。我没识别到连续的商品名称和数量数据区，请确认表头后面有实际订货数据。"
    return "Excel订单解析失败了。请确认文件是标准订单表，且表头后面包含商品和数量数据。"


def handle_excel_order_input(user_id: str, file_bytes: bytes, raw_ref: str) -> ChatResponse:
    started_at = time.perf_counter()
    try:
        payloads = parse_excel_order_payloads(file_bytes, raw_ref)
    except Exception as exc:
        logger.warning(
            "excel_order_import_failed user_id=%s raw_ref=%s size=%s signature=%s error=%s",
            user_id,
            raw_ref,
            len(file_bytes),
            excel_file_signature(file_bytes),
            exc,
            exc_info=True,
        )
        return ChatResponse(
            user_id=user_id,
            answer=excel_order_failure_reply(exc),
            history_length=user_order_count(user_id),
        )

    parsed_at = time.perf_counter()
    parsed_line_count = sum(len(payload.get("items") or []) for payload in payloads)
    logger.info(
        "excel_order_parse_done user_id=%s raw_ref=%s size=%s payloads=%s lines=%s elapsed_ms=%s",
        user_id,
        raw_ref,
        len(file_bytes),
        len(payloads),
        parsed_line_count,
        int((parsed_at - started_at) * 1000),
    )

    if len(payloads) != 1:
        logger.warning(
            "excel_order_multiple_payloads user_id=%s raw_ref=%s payload_count=%s elapsed_ms=%s",
            user_id,
            raw_ref,
            len(payloads),
            int((time.perf_counter() - started_at) * 1000),
        )
        return ChatResponse(
            user_id=user_id,
            answer="这份 Excel 里我解析出了多张订单。为避免记串，请按单张订单拆开发，或联系管理员处理。",
            history_length=user_order_count(user_id),
        )

    draft = normalize_order_draft(payloads[0])
    draft["confirmed"] = False
    draft["status"] = ORDER_STATUS_NEW
    draft["raw_ref"] = draft.get("raw_ref") or raw_ref
    save_started_at = time.perf_counter()
    save_order_draft(user_id, draft)
    save_ms = int((time.perf_counter() - save_started_at) * 1000)
    line_count = len(draft.get("items") or [])
    logger.info(
        "excel_order_draft_ready user_id=%s raw_ref=%s size=%s lines=%s parse_ms=%s draft_save_ms=%s elapsed_ms=%s",
        user_id,
        raw_ref,
        len(file_bytes),
        line_count,
        int((parsed_at - started_at) * 1000),
        save_ms,
        int((time.perf_counter() - started_at) * 1000),
    )
    summary = format_order_draft_summary(draft)
    return ChatResponse(
        user_id=user_id,
        answer=(
            f"我把 Excel 解析成待确认订单，共 {line_count} 行商品：\n"
            + summary
            + "\n确认无误请回复“确认 / 对 / ok / yes”；要修改就直接发修改内容。"
        ),
        history_length=user_order_count(user_id),
    )


def handle_photo_order_input(user_id: str, image_bytes: bytes, mime_type: str | None, raw_ref: str) -> ChatResponse:
    try:
        draft = llm_parse_photo_order(vision_client, VISION_MODEL, image_bytes, mime_type, raw_ref)
    except Exception as exc:
        logger.warning("photo_order_parse_failed user_id=%s raw_ref=%s error=%s", user_id, raw_ref, exc)
        if "vision model is not configured" in str(exc):
            return ChatResponse(
                user_id=user_id,
                answer="照片订单识别已收到，但当前视觉模型还没配置好。请先用文字发送门店、商品和数量。",
                history_length=user_order_count(user_id),
            )
        return ChatResponse(
            user_id=user_id,
            answer="这张照片我没有识别成功。请重新拍清楚订单表，或直接用文字发送门店、商品和数量。",
            history_length=user_order_count(user_id),
        )

    draft["confirmed"] = False
    draft["status"] = ORDER_STATUS_NEW
    save_order_draft(user_id, draft)

    missing = order_draft_missing_fields(draft)
    summary = format_order_draft_summary(draft)
    if missing:
        answer = (
            "我先把照片识别成订单草稿，还缺："
            + "、".join(missing)
            + "\n"
            + summary
            + "\n请直接补充缺失信息，或发“取消”清空。"
        )
    else:
        answer = (
            "我把照片识别成待确认订单：\n"
            + summary
            + "\n确认无误请回复“确认 / 对 / ok / yes”；要修改就直接发修改内容。"
        )

    return ChatResponse(
        user_id=user_id,
        answer=answer,
        history_length=user_order_count(user_id),
    )


def handle_wecom_kf_sync_item(item: dict[str, Any]) -> None:
    msg_id = str(item.get("msgid") or "")
    if is_duplicate_wecom_kf_message(msg_id):
        logger.info("wecom_kf_duplicate_message msg_id=%s", msg_id)
        return

    msg_type = item.get("msgtype")
    open_kfid = str(item.get("open_kfid") or item.get("event", {}).get("open_kfid") or "")
    external_userid = str(item.get("external_userid") or item.get("event", {}).get("external_userid") or "")

    logger.info(
        "wecom_kf_sync_item msg_id=%s msg_type=%s open_kfid=%s external_userid=%s origin=%s",
        msg_id,
        msg_type,
        open_kfid,
        external_userid,
        item.get("origin"),
    )

    if msg_type == "event":
        logger.info("wecom_kf_event msg_id=%s event=%s", msg_id, item.get("event", {}))
        return

    session_id = f"kf:{open_kfid}:{external_userid}" if open_kfid and external_userid else ""
    raw_ref = f"kf:{open_kfid}:{external_userid}:{msg_id}"

    if msg_type == "image":
        media_id = str(item.get("image", {}).get("media_id") or "")
        if not media_id or not open_kfid or not external_userid:
            logger.info("wecom_kf_image_skipped msg_id=%s", msg_id)
            return
        try:
            media_bytes, content_type, _filename = get_wecom_kf_media(media_id)
            if get_session_mode(session_id) == SESSION_MODE_RECEIPT:
                answer = handle_receipt_photo_input(session_id, media_bytes, content_type, raw_ref).answer
            else:
                answer = handle_photo_order_input(session_id, media_bytes, content_type, raw_ref).answer
        except Exception as exc:
            logger.exception("wecom_kf_image_order_failed msg_id=%s media_id=%s error=%s", msg_id, media_id, exc)
            answer = "这张图片处理失败了，请稍后再试，或直接用文字发送门店、商品和数量。"
        send_wecom_kf_text(open_kfid, external_userid, answer)
        return

    if msg_type == "file":
        media_id = str(item.get("file", {}).get("media_id") or "")
        filename = str(item.get("file", {}).get("filename") or item.get("file", {}).get("file_name") or media_id)
        if not media_id or not open_kfid or not external_userid:
            logger.info("wecom_kf_file_skipped msg_id=%s", msg_id)
            return
        try:
            media_bytes, content_type, downloaded_name = get_wecom_kf_media(media_id)
            filename = downloaded_name or filename
            extension = Path(filename).suffix.lower()
            logger.info(
                "wecom_kf_file_received msg_id=%s filename=%s content_type=%s size=%s signature=%s",
                msg_id,
                filename,
                content_type,
                len(media_bytes),
                excel_file_signature(media_bytes),
            )
            if extension not in {".xlsx", ".xlsm"} and "spreadsheet" not in content_type:
                answer = "这个文件我暂时只支持标准 Excel 订单表。"
            else:
                try:
                    send_wecom_kf_text(open_kfid, external_userid, "已收到Excel，正在解析订单内容，稍等一下。")
                except Exception as send_exc:
                    logger.warning(
                        "wecom_kf_file_processing_notice_failed msg_id=%s error=%s",
                        msg_id,
                        send_exc,
                    )
                answer = handle_excel_order_input(session_id, media_bytes, raw_ref=f"{raw_ref}:{filename}").answer
        except Exception as exc:
            logger.exception("wecom_kf_file_order_failed msg_id=%s media_id=%s error=%s", msg_id, media_id, exc)
            answer = "这个文件处理失败了。请确认是标准 Excel 订单表后重发。"
        send_wecom_kf_text(open_kfid, external_userid, answer)
        return

    if msg_type != "text":
        if open_kfid and external_userid:
            send_wecom_kf_text(open_kfid, external_userid, "目前我支持文字加单、订单照片和标准 Excel。")
        return

    content = str(item.get("text", {}).get("content") or "").strip()
    if not content or not open_kfid or not external_userid:
        logger.info("wecom_kf_text_skipped msg_id=%s", msg_id)
        return

    answer = handle_user_message(session_id, content, raw_ref=raw_ref).answer
    send_wecom_kf_text(open_kfid, external_userid, answer)


def sync_wecom_kf_messages(event: WecomKfEvent) -> None:
    if not event.token or not event.open_kfid:
        logger.warning("wecom_kf_event_missing_token_or_open_kfid event=%s", event.model_dump())
        return

    cursor = get_kf_cursor(event.open_kfid)

    for _ in range(20):
        payload = {
            "cursor": cursor,
            "token": event.token,
            "limit": WECOM_KF_SYNC_LIMIT,
            "open_kfid": event.open_kfid,
        }
        data = post_wecom_kf_api("kf/sync_msg", payload)
        next_cursor = str(data.get("next_cursor") or "")
        if next_cursor:
            set_kf_cursor(event.open_kfid, next_cursor)
            cursor = next_cursor

        for item in data.get("msg_list", []):
            if isinstance(item, dict):
                handle_wecom_kf_sync_item(item)

        if int(data.get("has_more", 0)) != 1:
            return

    logger.warning("wecom_kf_sync_stopped_after_page_limit open_kfid=%s", event.open_kfid)


def process_wecom_kf_event(event: WecomKfEvent) -> None:
    try:
        sync_wecom_kf_messages(event)
    except Exception as exc:
        logger.exception("wecom_kf_process_failed open_kfid=%s error=%s", event.open_kfid, exc)


def request_client_host(request: Request) -> str:
    return request.client.host if request.client else ""


def request_query_keys(request: Request) -> str:
    return ",".join(sorted(request.query_params.keys()))


@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok", "build": APP_BUILD_LABEL}


@app.get("/wecom/callback")
async def wecom_verify(request: Request):
    logger.info(
        "wecom_verify_hit client=%s query_keys=%s",
        request_client_host(request),
        request_query_keys(request),
    )
    msg_signature = require_query_param(request, "msg_signature")
    timestamp = require_query_param(request, "timestamp")
    nonce = require_query_param(request, "nonce")
    echostr = require_query_param(request, "echostr")

    ret, decrypted_echo = get_wecom_crypto().VerifyURL(
        msg_signature,
        timestamp,
        nonce,
        echostr,
    )
    if ret != 0:
        logger.warning("wecom_verify_failed ret=%s", ret)
        raise HTTPException(status_code=403, detail=f"WeCom verify failed: {ret}")

    if isinstance(decrypted_echo, bytes):
        decrypted_echo = decrypted_echo.decode("utf-8")

    return PlainTextResponse(decrypted_echo)


@app.post("/wecom/callback")
async def wecom_callback(request: Request):
    encrypted_body = await request.body()
    logger.info(
        "wecom_callback_hit client=%s query_keys=%s body_bytes=%s",
        request_client_host(request),
        request_query_keys(request),
        len(encrypted_body),
    )
    msg_signature = require_query_param(request, "msg_signature")
    timestamp = require_query_param(request, "timestamp")
    nonce = require_query_param(request, "nonce")

    ret, plain_xml = get_wecom_crypto().DecryptMsg(
        encrypted_body,
        msg_signature,
        timestamp,
        nonce,
    )
    if ret != 0:
        logger.warning("wecom_decrypt_failed ret=%s", ret)
        raise HTTPException(status_code=403, detail=f"WeCom decrypt failed: {ret}")

    message = parse_wecom_plain_message(plain_xml)
    logger.info(
        "wecom_message msg_id=%s msg_type=%s chat_type=%s chat_id=%s sender=%s",
        message.msg_id,
        message.msg_type,
        message.chat_type,
        message.chat_id,
        message.sender_user_id,
    )

    if is_duplicate_wecom_message(message.msg_id):
        logger.info("wecom_duplicate_message msg_id=%s", message.msg_id)
        answer = ""
    else:
        try:
            answer = answer_wecom_message(message)
        except Exception as exc:
            logger.warning("wecom_answer_failed msg_id=%s error=%s", message.msg_id, exc)
            answer = "这条消息处理失败了，请稍后再试。"

    encrypted_response = encrypt_wecom_response(answer, nonce, timestamp)
    return PlainTextResponse(encrypted_response, media_type="application/xml")


@app.get("/wecom/kf/callback")
async def wecom_kf_verify(request: Request):
    logger.info(
        "wecom_kf_verify_hit client=%s query_keys=%s",
        request_client_host(request),
        request_query_keys(request),
    )
    msg_signature = require_query_param(request, "msg_signature")
    timestamp = require_query_param(request, "timestamp")
    nonce = require_query_param(request, "nonce")
    echostr = require_query_param(request, "echostr")

    decrypted_echo = decrypt_wecom_kf_verify_url(
        msg_signature,
        timestamp,
        nonce,
        echostr,
    )

    return PlainTextResponse(decrypted_echo)


@app.post("/wecom/kf/callback")
async def wecom_kf_callback(request: Request, background_tasks: BackgroundTasks):
    encrypted_body = await request.body()
    logger.info(
        "wecom_kf_callback_hit client=%s query_keys=%s body_bytes=%s",
        request_client_host(request),
        request_query_keys(request),
        len(encrypted_body),
    )
    msg_signature = require_query_param(request, "msg_signature")
    timestamp = require_query_param(request, "timestamp")
    nonce = require_query_param(request, "nonce")

    plain_xml = decrypt_wecom_kf_message(
        encrypted_body,
        msg_signature,
        timestamp,
        nonce,
    )

    try:
        event = parse_wecom_kf_event(plain_xml)
    except Exception as exc:
        logger.warning("wecom_kf_parse_failed error=%s", exc)
        return PlainTextResponse("success")

    logger.info(
        "wecom_kf_callback event=%s open_kfid=%s create_time=%s",
        event.event,
        event.open_kfid,
        event.create_time,
    )
    background_tasks.add_task(process_wecom_kf_event, event)
    return PlainTextResponse("success")


@app.post("/chat", response_model=ChatResponse)
def chat(payload: ChatRequest) -> ChatResponse:
    return handle_user_message(payload.user_id, payload.message)


def parse_ids_param(ids: str | None) -> list[int]:
    if not ids:
        return []
    parsed: list[int] = []
    for part in ids.split(","):
        part = part.strip()
        if not part:
            continue
        try:
            parsed.append(int(part))
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"invalid id: {part}") from exc
    return parsed


def validate_iso_date_param(value: str, name: str) -> str:
    if not value:
        raise HTTPException(status_code=400, detail=f"{name} is required")
    try:
        return datetime.strptime(value, "%Y-%m-%d").date().isoformat()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=f"{name} must be YYYY-MM-DD") from exc


def require_robot_api_token(request: Request) -> None:
    authorization = request.headers.get("authorization", "")
    expected = f"Bearer {ROBOT_API_TOKEN}" if ROBOT_API_TOKEN else ""
    if not expected or not hmac.compare_digest(authorization, expected):
        raise HTTPException(
            status_code=401,
            detail="invalid robot api token",
            headers={"WWW-Authenticate": "Bearer"},
        )


@app.get("/api/orders")
def api_orders(
    request: Request,
    status: str = Query(default=ORDER_STATUS_NEW),
    order_date: str | None = Query(default=None),
    ids: str | None = Query(default=None),
) -> dict[str, list[dict[str, Any]]]:
    require_robot_api_token(request)
    parsed_ids = parse_ids_param(ids)
    if parsed_ids:
        return {"orders": query_order_payloads(ids=parsed_ids)}

    if status not in ORDER_STATUSES:
        raise HTTPException(status_code=400, detail="status must be new, fetched, or all")
    normalized_order_date = validate_iso_date_param(order_date or "", "order_date")
    query_status = None if status == ORDER_STATUS_ALL else status
    return {"orders": query_order_payloads(status=query_status, order_date=normalized_order_date)}


@app.post("/api/orders/mark_fetched")
def api_mark_fetched(request: Request, payload: MarkFetchedRequest) -> dict[str, Any]:
    require_robot_api_token(request)
    return mark_order_payloads_fetched(payload.ids)


@app.post("/api/orders/unmark")
def api_unmark_orders(request: Request, payload: MarkFetchedRequest) -> dict[str, Any]:
    require_robot_api_token(request)
    return unmark_order_payloads(payload.ids)


@app.get("/api/receipts")
def api_receipts(
    request: Request,
    date: str = Query(...),
    status: str = Query(default=RECEIPT_STATUS_NEW),
) -> dict[str, list[dict[str, Any]]]:
    require_robot_api_token(request)
    if status not in RECEIPT_API_STATUSES:
        raise HTTPException(status_code=400, detail="status must be new, fetched, or all")
    normalized_date = validate_iso_date_param(date, "date")
    return {"receipts": query_receipt_payloads_by_status(normalized_date, status)}


@app.post("/api/receipts/mark_fetched")
def api_mark_receipts_fetched(request: Request, payload: IdsRequest) -> dict[str, Any]:
    require_robot_api_token(request)
    return mark_receipt_payloads_fetched(payload.ids)


@app.post("/api/receipts/unmark")
def api_unmark_receipts(request: Request, payload: IdsRequest) -> dict[str, Any]:
    require_robot_api_token(request)
    return unmark_receipt_payloads(payload.ids)


@app.post("/api/orders/import/excel")
async def api_import_excel(
    request: Request,
    file: UploadFile = File(...),
    raw_ref: str | None = Query(default=None),
) -> dict[str, Any]:
    require_robot_api_token(request)
    file_bytes = await file.read()
    reference = raw_ref or file.filename or "api:excel"
    try:
        saved = save_excel_order_payloads(file_bytes, reference)
    except Exception as exc:
        logger.warning("api_excel_import_failed raw_ref=%s error=%s", reference, exc)
        raise HTTPException(status_code=400, detail="Excel order import failed") from exc
    return {"orders": saved}


@app.post("/api/orders/import/photo")
async def api_import_photo(
    request: Request,
    file: UploadFile = File(...),
    user_id: str = Query(default="api"),
    confirm: bool = Query(default=False),
    raw_ref: str | None = Query(default=None),
) -> dict[str, Any]:
    require_robot_api_token(request)
    image_bytes = await file.read()
    reference = raw_ref or file.filename or "api:photo"
    mime_type = file.content_type or mimetypes.guess_type(file.filename or "")[0]
    try:
        draft = llm_parse_photo_order(vision_client, VISION_MODEL, image_bytes, mime_type, reference)
    except Exception as exc:
        logger.warning("api_photo_import_failed raw_ref=%s error=%s", reference, exc)
        if "vision model is not configured" in str(exc):
            raise HTTPException(status_code=503, detail="Photo recognition vision model is not configured") from exc
        raise HTTPException(status_code=400, detail="Photo order parse failed") from exc

    if confirm:
        missing = order_draft_missing_fields(draft)
        if missing:
            raise HTTPException(status_code=400, detail="missing fields: " + "、".join(missing))
        draft["confirmed"] = True
        saved = insert_order_payload(draft)
        return {"order": saved}

    save_order_draft(user_id, draft)
    return {"draft": draft, "message": "photo parsed; confirm before it is visible from /api/orders"}


@app.post("/api/orders/import/text")
def api_import_text(request: Request, payload: TextOrderImportRequest) -> dict[str, Any]:
    require_robot_api_token(request)
    raw_ref = payload.raw_ref or f"api:text:{payload.user_id}"
    draft = llm_order_draft_from_message({}, payload.message)
    if not draft:
        logger.warning("api_text_import_failed user_id=%s", payload.user_id)
        raise HTTPException(status_code=400, detail="Text order parse failed")

    draft["raw_ref"] = raw_ref
    draft["raw_text"] = draft.get("raw_text") or payload.message
    draft["confirmed"] = bool(payload.confirm)
    if payload.confirm:
        missing = order_draft_missing_fields(draft)
        if missing:
            raise HTTPException(status_code=400, detail="missing fields: " + "、".join(missing))
        saved = insert_order_payload(draft)
        return {"order": saved}

    save_order_draft(payload.user_id, draft)
    return {"draft": normalize_order_payload(draft), "message": "text parsed; confirm before it is visible from /api/orders"}


@app.get("/memory/{user_id}", response_model=MemoryLengthResponse)
def get_memory_length(user_id: str) -> MemoryLengthResponse:
    user_id = user_id.strip()
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id cannot be empty")

    with MEMORY_LOCK:
        memory = load_memory()
        history = memory.get(user_id, [])

        if not isinstance(history, list):
            raise HTTPException(status_code=500, detail=f"Invalid history for user_id: {user_id}")

    return MemoryLengthResponse(user_id=user_id, history_length=len(history))


def require_export_token(request: Request) -> str:
    token = request.query_params.get("token", "")
    if EXPORT_TOKEN and token != EXPORT_TOKEN:
        raise HTTPException(status_code=403, detail="invalid export token")
    return token


@app.get("/exports", response_class=HTMLResponse)
def export_page(request: Request):
    token = request.query_params.get("token", "")

    if EXPORT_TOKEN and token != EXPORT_TOKEN:
        return HTMLResponse(
            """
            <!doctype html>
            <html lang="zh-CN">
              <head>
                <meta charset="utf-8">
                <meta name="viewport" content="width=device-width, initial-scale=1">
                <title>订单导出</title>
                <style>
                  body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 48px; color: #172033; }
                  main { max-width: 520px; }
                  label { display: block; margin: 18px 0 8px; font-weight: 600; }
                  input { width: 100%; box-sizing: border-box; padding: 12px 14px; border: 1px solid #ccd3df; border-radius: 8px; font-size: 16px; }
                  button { margin-top: 18px; padding: 12px 18px; border: 0; border-radius: 8px; background: #1f6feb; color: white; font-size: 16px; cursor: pointer; }
                  p { color: #667085; line-height: 1.6; }
                </style>
              </head>
              <body>
                <main>
                  <h1>订单导出</h1>
                  <p>请输入导出口令后进入导出页。</p>
                  <form method="get" action="/exports">
                    <label for="token">导出口令</label>
                    <input id="token" name="token" type="password" autocomplete="current-password">
                    <button type="submit">进入</button>
                  </form>
                </main>
              </body>
            </html>
            """
        )

    records = collect_order_records()
    order_ids = sorted({record["id"] for record in records if record.get("id")})
    stores = sorted({record["store"] for record in records if record.get("store")})
    download_url = "/exports/orders.xlsx"
    if token:
        download_url = f"{download_url}?token={token}"

    store_items = "".join(f"<li>{store}</li>" for store in stores[:30])
    if len(stores) > 30:
        store_items += f"<li>还有 {len(stores) - 30} 个门店...</li>"
    if not store_items:
        store_items = "<li>暂无门店数据</li>"

    return HTMLResponse(
        f"""
        <!doctype html>
        <html lang="zh-CN">
          <head>
            <meta charset="utf-8">
            <meta name="viewport" content="width=device-width, initial-scale=1">
            <title>订单导出</title>
            <style>
              body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", sans-serif; margin: 48px; color: #172033; background: #f6f8fb; }}
              main {{ max-width: 880px; }}
              .panel {{ background: white; border: 1px solid #e3e8f0; border-radius: 10px; padding: 28px; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.06); }}
              .stats {{ display: flex; gap: 16px; margin: 22px 0; }}
              .stat {{ background: #f1f5fb; border-radius: 8px; padding: 16px 18px; min-width: 140px; }}
              .stat strong {{ display: block; font-size: 28px; color: #0f4c81; }}
              a.button {{ display: inline-block; margin-top: 10px; padding: 13px 18px; border-radius: 8px; background: #1f6feb; color: white; text-decoration: none; font-weight: 700; }}
              p {{ color: #667085; line-height: 1.6; }}
              ul {{ columns: 2; color: #344054; line-height: 1.8; }}
            </style>
          </head>
          <body>
            <main class="panel">
              <h1>订单导出</h1>
              <p>点击按钮生成并下载 Excel。文件包含“全部订单”和“按门店商品汇总”两个 sheet。</p>
              <div class="stats">
                <div class="stat"><strong>{len(order_ids)}</strong>订单数</div>
                <div class="stat"><strong>{len(stores)}</strong>门店数</div>
              </div>
              <a class="button" href="{download_url}">下载 Excel</a>
              <h2>当前门店</h2>
              <ul>{store_items}</ul>
            </main>
          </body>
        </html>
        """
    )


@app.get("/exports/orders.xlsx")
def export_orders(request: Request):
    require_export_token(request)

    output_path = build_order_export(collect_order_records(), EXPORT_DIR)
    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=output_path.name,
    )


@app.delete("/memory/{user_id}", response_model=DeleteMemoryResponse)
def delete_memory(user_id: str) -> DeleteMemoryResponse:
    user_id = user_id.strip()
    if not user_id:
        raise HTTPException(status_code=400, detail="user_id cannot be empty")

    with MEMORY_LOCK:
        memory = load_memory()
        memory.pop(user_id, None)
        save_memory(memory)

    logger.info("memory_deleted user_id=%s", user_id)
    return DeleteMemoryResponse(deleted=True, user_id=user_id)
