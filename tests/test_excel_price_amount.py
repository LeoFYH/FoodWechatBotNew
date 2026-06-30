import unittest
from io import BytesIO

from openpyxl import Workbook

import excel_import


def _build_diananmen_order() -> bytes:
    """复刻"地安门"订货单表头：含税单价（元）与 金额（元）是两列。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "订单"
    ws["D1"] = "地安门"
    ws["D2"] = "2026年6月29日"
    ws["G2"] = "订货人："
    ws["D4"] = "商品名称"
    ws["E4"] = "商品规格"
    ws["F4"] = "单位"
    ws["G4"] = "含税单价（元）"
    ws["H4"] = "数量"
    ws["I4"] = "金额（元）"
    rows = [
        ("鸡汤鲜肉馄饨", "260g/袋*25袋", "箱", 267.32, 0.5, 133.66),
        ("老北京肉饼", "360g**5张", "袋", 44.62, 10, 446.20),
        ("鸡蛋面", None, "kg", 9.50, 50, 475.00),
        ("爆浆红糖饼", "330g*5个*12袋", "袋", 50.15, 5, 250.75),
    ]
    r = 5
    for name, spec, unit, price, qty, amount in rows:
        ws.cell(r, 4, name)
        ws.cell(r, 5, spec)
        ws.cell(r, 6, unit)
        ws.cell(r, 7, price)
        ws.cell(r, 8, qty)
        ws.cell(r, 9, amount)
        r += 1
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class HeaderKeyTests(unittest.TestCase):
    def test_price_headers(self):
        for header in ("含税单价（元）", "单价", "价格", "不含税单价"):
            self.assertEqual(excel_import.excel_header_key(header), "price", header)

    def test_amount_headers_are_not_price(self):
        for header in ("金额（元）", "金额", "总金额"):
            self.assertEqual(excel_import.excel_header_key(header), "amount", header)

    def test_spec_headers(self):
        for header in ("商品规格", "规格型号", "规格"):
            self.assertEqual(excel_import.excel_header_key(header), "spec", header)

    def test_qty_headers(self):
        for header in ("数量", "订货数量"):
            self.assertEqual(excel_import.excel_header_key(header), "qty", header)


class DiananmenParseTests(unittest.TestCase):
    def setUp(self):
        payloads = excel_import.parse_excel_order_payloads(_build_diananmen_order(), "test")
        self.assertEqual(len(payloads), 1)
        self.items = {it["name"]: it for it in payloads[0]["items"]}

    def test_price_is_unit_price_not_amount(self):
        # 核心回归：含税单价（G）进 price，金额（I）绝不进 price
        self.assertEqual(self.items["鸡汤鲜肉馄饨"]["price"], 267.32)
        self.assertEqual(self.items["老北京肉饼"]["price"], 44.62)
        self.assertEqual(self.items["鸡蛋面"]["price"], 9.5)
        self.assertEqual(self.items["爆浆红糖饼"]["price"], 50.15)

    def test_amount_captured_separately(self):
        self.assertEqual(self.items["鸡汤鲜肉馄饨"]["amount"], 133.66)
        self.assertEqual(self.items["老北京肉饼"]["amount"], 446.2)

    def test_amount_never_equals_price(self):
        for name, it in self.items.items():
            self.assertNotEqual(it["price"], it["amount"], name)

    def test_spec_and_qty(self):
        self.assertEqual(self.items["鸡汤鲜肉馄饨"]["spec"], "260g/袋*25袋")
        self.assertEqual(self.items["鸡汤鲜肉馄饨"]["qty"], 0.5)
        self.assertEqual(self.items["老北京肉饼"]["qty"], 10)


if __name__ == "__main__":
    unittest.main()
