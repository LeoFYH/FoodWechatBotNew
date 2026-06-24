import importlib
import os
import sys
import tempfile
import unittest
from io import BytesIO
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook


class ExcelParsingTests(unittest.TestCase):
    def setUp(self) -> None:
        self.tempdir = tempfile.TemporaryDirectory()
        root = Path(self.tempdir.name)
        self.env_patch = patch.dict(
            os.environ,
            {
                "LLM_API_KEY": "dummy",
                "DATABASE_BACKEND": "sqlite",
                "MEMORY_FILE": str(root / "memory.json"),
                "SESSION_STATE_FILE": str(root / "session_state.json"),
                "ORDER_DB_FILE": str(root / "orders.db"),
                "RECEIPT_DB_FILE": str(root / "receipts.db"),
                "INTERVIEW_ARCHIVE_FILE": str(root / "interviews.json"),
                "WECOM_KF_CURSOR_FILE": str(root / "kf_cursors.json"),
                "EXPORT_DIR": str(root / "exports"),
            },
            clear=False,
        )
        self.env_patch.start()
        os.environ.pop("VISION_API_KEY", None)
        sys.modules.pop("main", None)
        self.main = importlib.import_module("main")

    def tearDown(self) -> None:
        sys.modules.pop("main", None)
        self.env_patch.stop()
        self.tempdir.cleanup()

    def workbook_bytes(self, workbook: Workbook) -> bytes:
        output = BytesIO()
        workbook.save(output)
        return output.getvalue()

    def test_parse_flexible_headers_and_quantity_unit(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["收货门店", "鼓楼店", "订单日期", "2026-06-24"])
        sheet.append(["商品名称/规格", "订货数量(箱)", "单价"])
        sheet.append(["鲜肉馄饨 260g", 2, 128.5])

        payloads = self.main.parse_excel_order_payloads(self.workbook_bytes(workbook), "test.xlsx")

        self.assertEqual(len(payloads), 1)
        payload = payloads[0]
        self.assertEqual(payload["store"], "鼓楼店")
        self.assertEqual(payload["order_date"], "2026-06-24")
        self.assertEqual(payload["items"][0]["name"], "鲜肉馄饨 260g")
        self.assertEqual(payload["items"][0]["qty"], 2)
        self.assertEqual(payload["items"][0]["unit"], "箱")
        self.assertEqual(payload["items"][0]["price"], 128.5)

    def test_parse_order_sheet_when_active_sheet_is_not_order(self) -> None:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "说明"
        summary_sheet.append(["这个 sheet 不是订单"])
        order_sheet = workbook.create_sheet("订单")
        order_sheet.append(["门店名称", "订单日期"])
        order_sheet.append(["商品名称（规格）", "数量（袋）"])
        order_sheet.append(["虾仁馄饨", 6])

        payloads = self.main.parse_excel_order_payloads(self.workbook_bytes(workbook), "multi.xlsx")

        self.assertEqual(len(payloads), 1)
        self.assertEqual(payloads[0]["store"], "未确认门店")
        self.assertEqual(payloads[0]["items"][0]["name"], "虾仁馄饨")
        self.assertEqual(payloads[0]["items"][0]["qty"], 6)
        self.assertEqual(payloads[0]["items"][0]["unit"], "袋")

    def test_parse_unknown_headers_by_following_data_rows(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["门店", "鼓楼店"])
        sheet.append(["货品信息", "采购数"])
        sheet.append(["香菇馄饨", 3])
        sheet.append(["牛肉馄饨", 5])

        payloads = self.main.parse_excel_order_payloads(self.workbook_bytes(workbook), "fuzzy.xlsx")

        self.assertEqual(len(payloads), 1)
        self.assertEqual(payloads[0]["store"], "鼓楼店")
        self.assertEqual([item["name"] for item in payloads[0]["items"]], ["香菇馄饨", "牛肉馄饨"])
        self.assertEqual([item["qty"] for item in payloads[0]["items"]], [3, 5])

    def test_reject_non_excel_download_content(self) -> None:
        with self.assertRaisesRegex(ValueError, "not Excel"):
            self.main.parse_excel_order_payloads(b'{"errcode":40007}', "media.xlsx")

    def test_skip_header_like_sheet_without_item_rows(self) -> None:
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "汇总"
        summary_sheet.append(["商品名称", "订货数量"])
        summary_sheet.append(["说明", None])
        order_sheet = workbook.create_sheet("鼓楼订货")
        order_sheet.append(["非标准品项", "要货数"])
        order_sheet.append(["虾仁馄饨", 6])

        payloads = self.main.parse_excel_order_payloads(self.workbook_bytes(workbook), "skip-empty.xlsx")

        self.assertEqual(len(payloads), 1)
        self.assertEqual(payloads[0]["items"][0]["name"], "虾仁馄饨")
        self.assertEqual(payloads[0]["items"][0]["qty"], 6)

    def test_parse_catalog_order_form_with_sparse_quantities(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "鼓楼"
        sheet.append(["馄饨侯（鼓楼）店产品订货单"])
        sheet.append(["订货日期：", None, None, 46193, None, None, None, "订货人：", "周凯"])
        sheet.append(["到货日期：", None, None, 46194])
        sheet.append(["序号", "类别", None, "原料名称", "规格", "单位", "单价", "订货数量"])
        sheet.append([1, "馄饨", "05020093", "鸡汤鲜肉馄饨", "260g/袋*25袋", "箱", 267.32, None])
        sheet.append([2, "馄饨", "05020094", "鸡汤虾肉馄饨", "500g/袋*12袋", "箱", 399.11, 1])
        sheet.append([3, "面条类", "03010001", "碱面条", None, "kg", 11.96, 4])

        payloads = self.main.parse_excel_order_payloads(self.workbook_bytes(workbook), "catalog.xlsx")

        self.assertEqual(len(payloads), 1)
        payload = payloads[0]
        self.assertEqual(payload["store"], "鼓楼")
        self.assertEqual(payload["order_date"], "2026-06-20")
        self.assertEqual(payload["deliver_date"], "2026-06-21")
        self.assertEqual(len(payload["items"]), 2)
        self.assertEqual(payload["items"][0]["code"], "05020094")
        self.assertEqual(payload["items"][0]["name"], "鸡汤虾肉馄饨")
        self.assertEqual(payload["items"][1]["code"], "03010001")
        self.assertEqual(payload["items"][1]["qty"], 4)

    def test_excel_wechat_input_creates_confirmation_draft(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["收货门店", "鼓楼店", "订单日期", "2026-06-24"])
        sheet.append(["商品名称", "订货数量(箱)"])
        sheet.append(["鲜肉馄饨", 2])

        response = self.main.handle_excel_order_input(
            "u1",
            self.workbook_bytes(workbook),
            "kf:file:test.xlsx",
        )

        draft = self.main.get_order_draft("u1")
        self.assertIn("待确认订单", response.answer)
        self.assertIn("确认", response.answer)
        self.assertEqual(self.main.get_session_mode("u1"), self.main.SESSION_MODE_ORDER)
        self.assertEqual(draft["source"], self.main.ORDER_SOURCE_EXCEL)
        self.assertFalse(draft["confirmed"])
        self.assertEqual(draft["store"], "鼓楼店")
        self.assertEqual(draft["items"][0]["name"], "鲜肉馄饨")
        self.assertEqual(draft["items"][0]["qty"], 2)
        self.assertEqual(self.main.query_order_payloads(), [])

    def test_wecom_file_sends_processing_message_before_excel_result(self) -> None:
        sent_messages: list[str] = []

        with patch.object(self.main, "is_duplicate_wecom_kf_message", return_value=False), \
            patch.object(
                self.main,
                "get_wecom_kf_media",
                return_value=(b"fake-xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "test.xlsx"),
            ), \
            patch.object(
                self.main,
                "handle_excel_order_input",
                return_value=self.main.ChatResponse(user_id="u1", answer="解析完成", history_length=0),
            ), \
            patch.object(self.main, "send_wecom_kf_text", side_effect=lambda _kf, _user, content: sent_messages.append(content)):
            self.main.handle_wecom_kf_sync_item(
                {
                    "msgid": "m1",
                    "msgtype": "file",
                    "open_kfid": "kf1",
                    "external_userid": "user1",
                    "file": {"media_id": "media1", "filename": "test.xlsx"},
                }
            )

        self.assertGreaterEqual(len(sent_messages), 2)
        self.assertIn("已收到Excel", sent_messages[0])
        self.assertIn("正在解析", sent_messages[0])
        self.assertEqual(sent_messages[-1], "解析完成")


if __name__ == "__main__":
    unittest.main()
