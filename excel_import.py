"""excel_import.py —— Excel 订单解析层（从 main.py 原样搬出）。

把上传的 .xlsx 订货单识别成统一的订单 payload 列表：找表头、抽元数据、
逐行归一化成 order payload。全是纯解析逻辑：依赖标准库 + openpyxl + order_normalize
的归一化原语，绝不 import main，无 env/锁/client 依赖，可被 main.py 通过
`from excel_import import *` 门面 re-export，调用点与测试可见性不变。

注意：score_excel_header_candidate 与 find_excel_header_row 必须同处一模块
（find 内部调用 score）。测试若要 patch score，需 patch 本模块的名字（excel_import.score_*）。
"""

from __future__ import annotations

import logging
import re
import time
from datetime import date, datetime
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

from order_normalize import (
    ORDER_KIND_BASE,
    ORDER_SOURCE_EXCEL,
    ORDER_STATUS_NEW,
    clean_export_value,
    clean_order_value,
    normalize_base_item,
    normalize_date_text,
    normalize_number,
    normalize_order_date_text,
    normalize_order_payload,
    now_iso,
)

logger = logging.getLogger("wechatclaw")

EXCEL_MAX_SCAN_ROWS = 500
EXCEL_MAX_SCAN_COLUMNS = 80

EXCEL_HEADER_ALIASES = {
    "store": {"门店", "门店/区域", "区域", "店铺", "店名", "客户", "客户名称", "收货方", "门店名称", "收货门店", "收货店铺"},
    "order_no": {"订单号", "单号", "订单编号", "编号"},
    "orderer": {"下单人", "订货人", "订货员", "制单人", "联系人"},
    "order_date": {"下单日期", "订单日期", "订货日期", "日期", "制单日期"},
    "deliver_date": {"送达日期", "送货日期", "配送日期", "交付日期", "到货日期"},
    "code": {"商品编码", "编码", "货号", "商品代码", "code", "物料编码"},
    "name": {"商品名称", "商品名称规格", "品名规格", "品名", "名称", "商品", "产品名称", "货品名称", "货物名称", "name", "物料名称", "原料名称"},
    "spec": {"规格", "规格型号", "型号", "包装规格", "spec"},
    "unit": {"单位", "unit"},
    "qty": {"数量", "订货数量", "订购数量", "下单数量", "箱数", "件数", "qty"},
    "price": {"单价", "价格", "price"},
    "category": {"分类", "类别", "品类", "category"},
}

EXCEL_METADATA_LABELS = {
    "store": {"门店", "门店/区域", "区域", "店铺", "客户", "收货方", "门店名称", "收货门店", "收货店铺"},
    "order_no": {"订单号", "单号", "订单编号"},
    "orderer": {"下单人", "订货人", "联系人"},
    "order_date": {"下单日期", "订单日期", "订货日期"},
    "deliver_date": {"送达日期", "送货日期", "配送日期", "到货日期"},
}


def normalize_excel_header(value: Any) -> str:
    text = clean_order_value(value).lower()
    return re.sub(r"[\s:_：/\\（）()\[\]【】\-]+", "", text)


def excel_header_key(value: Any) -> str | None:
    normalized = normalize_excel_header(value)
    if not normalized:
        return None
    for key, aliases in EXCEL_HEADER_ALIASES.items():
        for alias in aliases:
            if normalized == normalize_excel_header(alias):
                return key

    if normalized in {"商品名称规格型号", "产品名称规格", "货品名称规格", "物料名称规格"}:
        return "name"
    if normalized.startswith("商品名称") and "编码" not in normalized and "代码" not in normalized:
        return "name"
    if normalized.startswith("品名") or normalized.startswith("产品名称") or normalized.startswith("货品名称"):
        return "name"
    if normalized.endswith("门店") or normalized.endswith("店铺"):
        return "store"
    if normalized.startswith("订货数量") or normalized.startswith("订购数量") or normalized.startswith("下单数量"):
        return "qty"
    if normalized.startswith("数量") and not normalized.startswith("数量合计"):
        return "qty"
    if normalized.endswith("数量") and not normalized.endswith("订单数量"):
        return "qty"
    return None


def excel_qty_unit_from_header(value: Any) -> str:
    text = clean_order_value(value)
    if not text:
        return ""
    bracket_match = re.search(r"[（(]\s*([^）)]+?)\s*[）)]", text)
    if bracket_match:
        unit = clean_order_value(bracket_match.group(1)).strip()
        if unit and len(unit) <= 8:
            return unit
    normalized = normalize_excel_header(text)
    for unit in ("箱", "件", "袋", "盒", "包", "斤", "公斤", "kg", "KG", "份", "个", "瓶", "桶", "条", "只"):
        if normalized.endswith(unit.lower()) or normalized.endswith(unit):
            return unit
    return ""


def excel_header_units(row: tuple[Any, ...], header_map: dict[int, str]) -> dict[str, str]:
    units: dict[str, str] = {}
    for index, key in header_map.items():
        if key == "qty" and index < len(row):
            unit = excel_qty_unit_from_header(row[index])
            if unit:
                units[key] = unit
    return units


def excel_file_signature(file_bytes: bytes) -> str:
    sample = file_bytes[:64].lstrip()
    if not sample:
        return "empty"
    if sample.startswith(b"PK\x03\x04"):
        return "xlsx_zip"
    if sample.startswith(b"\xd0\xcf\x11\xe0"):
        return "legacy_xls"
    if sample.startswith(b"{") or sample.startswith(b"["):
        return "json"
    if sample[:16].lower().startswith((b"<!doctype html", b"<html", b"<?xml")):
        return "text_markup"
    return "unknown"


def ensure_excel_file_content(file_bytes: bytes) -> None:
    signature = excel_file_signature(file_bytes)
    if signature == "xlsx_zip":
        return
    if signature == "legacy_xls":
        raise ValueError("Excel content is legacy .xls; please send .xlsx")
    if signature in {"empty", "json", "text_markup"}:
        raise ValueError(f"Downloaded content is not Excel: {signature}")
    raise ValueError("Downloaded content is not a valid .xlsx file")


def excel_row_has_value(row: tuple[Any, ...]) -> bool:
    return any(clean_order_value(value) for value in row)


def excel_cell_value(row: tuple[Any, ...], index: int) -> Any:
    return row[index] if index < len(row) else None


def excel_label_header_map(row: tuple[Any, ...]) -> dict[int, str]:
    header_map: dict[int, str] = {}
    for column_index, value in enumerate(row):
        key = excel_header_key(value)
        if key and key not in header_map.values():
            header_map[column_index] = key
    return header_map


def is_excel_date_like_value(value: Any) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    return isinstance(value, (int, float)) and not isinstance(value, bool) and 20000 <= float(value) <= 80000


def is_excel_unit_text(value: Any) -> bool:
    text = clean_order_value(value)
    if not text or len(text) > 8:
        return False
    normalized = normalize_excel_header(text)
    return normalized in {
        "箱",
        "件",
        "袋",
        "盒",
        "包",
        "斤",
        "公斤",
        "kg",
        "份",
        "个",
        "瓶",
        "桶",
        "条",
        "只",
        "套",
    }


def is_excel_summary_name(value: Any) -> bool:
    normalized = normalize_excel_header(value)
    return normalized in {"合计", "小计", "总计", "共计", "合计数量", "数量合计"}


def looks_like_excel_item_code(value: Any) -> bool:
    text = clean_order_value(value)
    if not text or text == "#N/A":
        return bool(text)
    compact = re.sub(r"\s+", "", text)
    if len(compact) < 4:
        return False
    if not re.search(r"\d", compact):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9#/_-]+", compact))


def looks_like_excel_item_name(value: Any) -> bool:
    text = clean_order_value(value)
    if not text or len(re.sub(r"\s+", "", text)) < 2:
        return False
    if is_excel_summary_name(text) or is_excel_metadata_label(text):
        return False
    if excel_header_key(text) or looks_like_excel_item_code(text) or is_excel_unit_text(text):
        return False
    if normalize_number(text) is not None and re.fullmatch(r"[-+]?[\d,]+(?:\.\d+)?", text):
        return False
    return True


def excel_quantity_number(value: Any) -> int | float | None:
    if is_excel_date_like_value(value):
        return None
    if looks_like_excel_item_code(value):
        return None
    return normalize_number(value)


def infer_excel_code_column(rows: list[tuple[Any, ...]], header_index: int, header_map: dict[int, str]) -> int | None:
    if "code" in header_map.values() or "name" not in header_map.values():
        return None
    name_index = next(index for index, key in header_map.items() if key == "name")
    candidates = [
        index
        for index in range(max(0, name_index - 3), name_index)
        if index not in header_map
    ]
    best_index: int | None = None
    best_score = 0
    for index in candidates:
        score = 0
        for row in rows[header_index + 1 : header_index + 25]:
            if index < len(row) and looks_like_excel_item_code(row[index]):
                score += 1
        if score > best_score:
            best_score = score
            best_index = index
    return best_index if best_score >= 2 else None


def excel_candidate_columns(row: tuple[Any, ...], data_rows: list[tuple[Any, ...]]) -> list[int]:
    candidates: set[int] = set()
    for candidate_row in [row, *data_rows[:80]]:
        for index, value in enumerate(candidate_row):
            if clean_order_value(value):
                candidates.add(index)
    return sorted(candidates)


def score_excel_header_candidate(rows: list[tuple[Any, ...]], header_index: int) -> tuple[int, dict[int, str], int]:
    header_row = rows[header_index]
    label_map = excel_label_header_map(header_row)
    data_rows = [row for row in rows[header_index + 1 :] if excel_row_has_value(row)]
    if not data_rows:
        return 0, {}, 0

    bad_name_keys = {"store", "order_no", "orderer", "order_date", "deliver_date", "code", "unit", "qty", "price", "category"}
    bad_qty_keys = {"store", "order_no", "orderer", "order_date", "deliver_date", "code", "name", "spec", "unit", "price", "category"}
    columns = list(excel_candidate_columns(header_row, data_rows))
    best_score = 0
    best_map: dict[int, str] = {}
    best_item_count = 0

    for name_index in columns:
        name_key = label_map.get(name_index)
        if name_key in bad_name_keys:
            continue
        for qty_index in columns:
            if qty_index == name_index:
                continue
            qty_key = label_map.get(qty_index)
            if qty_key in bad_qty_keys:
                continue

            hits: list[tuple[int, str, int | float]] = []
            integer_qty_count = 0
            decimal_qty_count = 0
            blank_qty_count = 0
            for offset, data_row in enumerate(data_rows):
                qty_number = excel_quantity_number(excel_cell_value(data_row, qty_index))
                name_value = excel_cell_value(data_row, name_index)
                if qty_number is None:
                    if looks_like_excel_item_name(name_value):
                        blank_qty_count += 1
                    continue
                if not looks_like_excel_item_name(name_value):
                    continue
                name_text = clean_order_value(name_value)
                hits.append((offset, name_text, qty_number))
                if isinstance(qty_number, float) and not qty_number.is_integer():
                    decimal_qty_count += 1
                else:
                    integer_qty_count += 1

            if not hits:
                continue

            first_hit_offset = hits[0][0]
            immediate_hit_count = sum(1 for offset, _, _ in hits if offset <= 5)
            if immediate_hit_count == 0 and len(hits) < 2:
                continue

            unique_names = len({name for _, name, _ in hits})
            avg_name_length = sum(len(name) for _, name, _ in hits) / len(hits)
            score = len(hits) * 20
            score += immediate_hit_count * 8
            score += min(unique_names, 8) * 3
            score += min(int(avg_name_length), 12)
            score += min(integer_qty_count, 8) * 2
            score += min(blank_qty_count, 8)
            if qty_index > name_index:
                score += 6
            if label_map.get(name_index) == "name":
                score += 40
            if label_map.get(qty_index) == "qty":
                score += 45
            score += len(label_map) * 2
            score -= first_hit_offset * 3
            score -= decimal_qty_count * 3

            if score > best_score:
                candidate_map = dict(label_map)
                candidate_map[name_index] = "name"
                candidate_map[qty_index] = "qty"
                best_score = score
                best_map = candidate_map
                best_item_count = len(hits)

    if best_score < 24:
        return 0, {}, 0
    return best_score, best_map, best_item_count


def count_labeled_excel_items(rows: list[tuple[Any, ...]], header_index: int, header_map: dict[int, str]) -> int:
    name_indexes = [index for index, key in header_map.items() if key == "name"]
    qty_indexes = [index for index, key in header_map.items() if key == "qty"]
    if not name_indexes or not qty_indexes:
        return 0

    item_count = 0
    for row in rows[header_index + 1 : header_index + 160]:
        if not excel_row_has_value(row):
            continue
        for name_index in name_indexes:
            name_value = excel_cell_value(row, name_index)
            if not looks_like_excel_item_name(name_value):
                continue
            if any(excel_quantity_number(excel_cell_value(row, qty_index)) is not None for qty_index in qty_indexes):
                item_count += 1
                break
    return item_count


def find_excel_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[int, str]]:
    has_labeled_header = False
    for index, row in enumerate(rows):
        if not excel_row_has_value(row):
            continue
        header_map = excel_label_header_map(row)
        if "name" not in header_map.values() or "qty" not in header_map.values():
            continue
        has_labeled_header = True
        if count_labeled_excel_items(rows, index, header_map) > 0:
            return index, header_map
    if has_labeled_header:
        raise ValueError("Excel file contains no order item rows after labeled header")

    best_index = -1
    best_map: dict[int, str] = {}
    best_score = 0
    best_item_count = 0
    for index, row in enumerate(rows):
        if not excel_row_has_value(row):
            continue
        score, header_map, item_count = score_excel_header_candidate(rows, index)
        if score > best_score:
            best_index = index
            best_map = header_map
            best_score = score
            best_item_count = item_count

    if best_index < 0 or not best_map or best_item_count < 1:
        raise ValueError("Excel header row not found; expected item rows after a header row")
    return best_index, best_map


def is_excel_metadata_label(value: Any) -> bool:
    value_text = normalize_excel_header(value)
    if not value_text:
        return False
    for labels in EXCEL_METADATA_LABELS.values():
        if any(value_text == normalize_excel_header(label) for label in labels):
            return True
    return False


def next_excel_metadata_value(cells: list[Any], index: int) -> Any:
    for value in cells[index + 1 :]:
        if not clean_order_value(value):
            continue
        if is_excel_metadata_label(value):
            return None
        return value
    return None


def extract_store_from_excel_title(text: str) -> str:
    cleaned = clean_order_value(text)
    patterns = [
        r"馄饨侯[（(]([^）)]+)[）)]店?产品?订货单",
        r"馄饨侯(.+?)店产品?订货单",
        r"(.+?)店产品?订货单",
    ]
    for pattern in patterns:
        match = re.search(pattern, cleaned)
        if match:
            return clean_export_value(match.group(1))
    return ""


def extract_excel_metadata(rows: list[tuple[Any, ...]], header_index: int) -> dict[str, Any]:
    metadata: dict[str, Any] = {}
    for row in rows[: max(header_index, 1)]:
        cells = list(row)
        for index, value in enumerate(cells):
            text = clean_order_value(value)
            if not text:
                continue
            if not metadata.get("store"):
                title_store = extract_store_from_excel_title(text)
                if title_store:
                    metadata["store"] = title_store
            for field, labels in EXCEL_METADATA_LABELS.items():
                if metadata.get(field):
                    continue
                for label in labels:
                    label_text = normalize_excel_header(label)
                    value_text = normalize_excel_header(text)
                    if value_text == label_text:
                        metadata[field] = next_excel_metadata_value(cells, index)
                    else:
                        inline_match = re.match(rf"^\s*{re.escape(label)}\s*[：:]\s*(.+?)\s*$", text)
                        if inline_match:
                            metadata[field] = inline_match.group(1)
    return metadata


def row_value_by_header(row: tuple[Any, ...], header_map: dict[int, str], field: str) -> Any:
    for index, key in header_map.items():
        if key == field and index < len(row):
            return row[index]
    return None


def finalize_excel_header_map(rows: list[tuple[Any, ...]], header_index: int, header_map: dict[int, str]) -> dict[int, str]:
    finalized = dict(header_map)
    inferred_code_column = infer_excel_code_column(rows, header_index, finalized)
    if inferred_code_column is not None:
        finalized[inferred_code_column] = "code"
    return finalized


def worksheet_value_rows(sheet: Any) -> list[tuple[Any, ...]]:
    cells = getattr(sheet, "_cells", None)
    if isinstance(cells, dict) and cells:
        valued_cells: list[tuple[int, int, Any]] = []
        for cell in cells.values():
            value = getattr(cell, "value", None)
            if value is None:
                continue
            row_index = getattr(cell, "row", None)
            column_index = getattr(cell, "column", None)
            if not isinstance(row_index, int) or not isinstance(column_index, int):
                continue
            valued_cells.append((row_index, column_index, value))

        if not valued_cells:
            return []

        min_row = min(row for row, _column, _value in valued_cells)
        max_row = max(row for row, _column, _value in valued_cells)
        min_column = min(column for _row, column, _value in valued_cells)
        max_column = max(column for _row, column, _value in valued_cells)
        rows: list[list[Any]] = [
            [None] * (max_column - min_column + 1)
            for _ in range(max_row - min_row + 1)
        ]
        for row_index, column_index, value in valued_cells:
            rows[row_index - min_row][column_index - min_column] = value
        return [tuple(row) for row in rows]

    max_row = min(int(getattr(sheet, "max_row", EXCEL_MAX_SCAN_ROWS) or EXCEL_MAX_SCAN_ROWS), EXCEL_MAX_SCAN_ROWS)
    max_column = min(int(getattr(sheet, "max_column", EXCEL_MAX_SCAN_COLUMNS) or EXCEL_MAX_SCAN_COLUMNS), EXCEL_MAX_SCAN_COLUMNS)
    return list(sheet.iter_rows(max_row=max_row, max_col=max_column, values_only=True))


def find_excel_order_tables(workbook: Any) -> list[tuple[str, list[tuple[Any, ...]], int, dict[int, str]]]:
    tables: list[tuple[str, list[tuple[Any, ...]], int, dict[int, str]]] = []
    for sheet in workbook.worksheets:
        rows = worksheet_value_rows(sheet)
        if not rows:
            continue
        try:
            header_index, header_map = find_excel_header_row(rows)
        except ValueError:
            continue
        tables.append((sheet.title, rows, header_index, finalize_excel_header_map(rows, header_index, header_map)))
    return tables


def find_excel_order_rows(workbook: Any) -> tuple[list[tuple[Any, ...]], int, dict[int, str]]:
    last_error: ValueError | None = None
    for sheet in workbook.worksheets:
        rows = worksheet_value_rows(sheet)
        if not rows:
            continue
        try:
            header_index, header_map = find_excel_header_row(rows)
        except ValueError as exc:
            last_error = exc
            continue
        return rows, header_index, finalize_excel_header_map(rows, header_index, header_map)
    if last_error:
        raise last_error
    raise ValueError("Excel file is empty")


def parse_excel_order_payloads(file_bytes: bytes, raw_ref: str) -> list[dict[str, Any]]:
    ensure_excel_file_content(file_bytes)
    started_at = time.perf_counter()
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    load_ms = int((time.perf_counter() - started_at) * 1000)
    logger.info(
        "excel_workbook_loaded raw_ref=%s size=%s sheets=%s elapsed_ms=%s",
        raw_ref,
        len(file_bytes),
        len(workbook.worksheets),
        load_ms,
    )
    scan_started_at = time.perf_counter()
    tables = find_excel_order_tables(workbook)
    if not tables:
        find_excel_order_rows(workbook)
    logger.info(
        "excel_order_tables_scanned raw_ref=%s sheets=%s tables=%s elapsed_ms=%s",
        raw_ref,
        len(workbook.worksheets),
        len(tables),
        int((time.perf_counter() - scan_started_at) * 1000),
    )
    grouped: dict[tuple[str, str, str, str, str], dict[str, Any]] = {}

    for sheet_title, rows, header_index, header_map in tables:
        metadata = extract_excel_metadata(rows, header_index)
        if not metadata.get("store"):
            metadata["store"] = extract_store_from_excel_title(sheet_title)
        header_units = excel_header_units(rows[header_index], header_map)

        for row in rows[header_index + 1 :]:
            item = {
                "code": row_value_by_header(row, header_map, "code"),
                "name": row_value_by_header(row, header_map, "name"),
                "spec": row_value_by_header(row, header_map, "spec"),
                "unit": row_value_by_header(row, header_map, "unit") or header_units.get("qty"),
                "qty": row_value_by_header(row, header_map, "qty"),
                "price": row_value_by_header(row, header_map, "price"),
                "category": row_value_by_header(row, header_map, "category"),
            }
            normalized_item = normalize_base_item(item)
            if not normalized_item.get("name") or normalized_item.get("qty") is None:
                continue
            if is_excel_summary_name(normalized_item.get("name")):
                continue

            store = row_value_by_header(row, header_map, "store") or metadata.get("store") or ""
            order_no = row_value_by_header(row, header_map, "order_no") or metadata.get("order_no") or ""
            orderer = row_value_by_header(row, header_map, "orderer") or metadata.get("orderer") or ""
            order_date = row_value_by_header(row, header_map, "order_date") or metadata.get("order_date") or ""
            deliver_date = row_value_by_header(row, header_map, "deliver_date") or metadata.get("deliver_date") or ""

            key = (
                clean_order_value(store),
                clean_order_value(order_no),
                clean_order_value(orderer),
                normalize_order_date_text(order_date),
                normalize_date_text(deliver_date),
            )
            payload = grouped.setdefault(
                key,
                {
                    "kind": ORDER_KIND_BASE,
                    "source": ORDER_SOURCE_EXCEL,
                    "store": store,
                    "order_no": order_no,
                    "orderer": orderer,
                    "order_date": order_date,
                    "deliver_date": deliver_date,
                    "items": [],
                    "confirmed": True,
                    "status": ORDER_STATUS_NEW,
                    "raw_ref": raw_ref,
                    "created_at": now_iso(),
                },
            )
            payload["items"].append(normalized_item)

    payloads = [normalize_order_payload(payload) for payload in grouped.values()]
    if not payloads:
        raise ValueError("Excel file contains no order item rows")
    return payloads


__all__ = [
    # Excel 专用常量
    "EXCEL_MAX_SCAN_ROWS",
    "EXCEL_MAX_SCAN_COLUMNS",
    "EXCEL_HEADER_ALIASES",
    "EXCEL_METADATA_LABELS",
    # 解析函数
    "normalize_excel_header",
    "excel_header_key",
    "excel_qty_unit_from_header",
    "excel_header_units",
    "excel_file_signature",
    "ensure_excel_file_content",
    "excel_row_has_value",
    "excel_cell_value",
    "excel_label_header_map",
    "is_excel_date_like_value",
    "is_excel_unit_text",
    "is_excel_summary_name",
    "looks_like_excel_item_code",
    "looks_like_excel_item_name",
    "excel_quantity_number",
    "infer_excel_code_column",
    "excel_candidate_columns",
    "score_excel_header_candidate",
    "count_labeled_excel_items",
    "find_excel_header_row",
    "is_excel_metadata_label",
    "next_excel_metadata_value",
    "extract_store_from_excel_title",
    "extract_excel_metadata",
    "row_value_by_header",
    "finalize_excel_header_map",
    "worksheet_value_rows",
    "find_excel_order_tables",
    "find_excel_order_rows",
    "parse_excel_order_payloads",
]
